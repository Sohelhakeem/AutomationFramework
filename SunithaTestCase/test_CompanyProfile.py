import os
import time

import pytest
from openpyxl.reader.excel import load_workbook
# import self
from selenium import webdriver
from selenium.webdriver import ActionChains, Keys
from selenium.webdriver.common.by import By
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC

from pageObjects.LoginPage import LoginPage
from sunithaPageObjects.CompanyProfile import LoginPage
from utilities.customLogger import LogGen
from pageObjects.randomGen import randomGen
from utilities.readProperties import ReadConfig
from sunithaPageObjects.MyProfile import MyprofilePage
from GenericLib.BaseClass import BaseClass

class TestCompanyProfile(BaseClass):
    baseURL = ReadConfig.getApplicationURL()
    # username = ReadConfig.getUseremail()
    password = ReadConfig.getPassword()
    relative_three = "Files/three.png"
    absolute_path3 = os.path.abspath(relative_three)
    relative_four = "Files/four.png"
    absolute_path4 = os.path.abspath(relative_four)
    relative_five = "Files/five.png"
    absolute_path5 = os.path.abspath(relative_five)
    relative_six = "Files/six.jpg"
    absolute_path6 = os.path.abspath(relative_six)
    relative_seven = "Files/seven.jpg"
    absolute_path7 = os.path.abspath(relative_seven)

    workbook = load_workbook("TestData/LoginData.xlsx")

    # Access the active worksheet
    worksheet = workbook.active

    username = worksheet["A2"].value

    workbook.close()
    # Load the existing workbook
    wb = load_workbook("TestData/LoginData.xlsx")

    # Select the active worksheet
    ws = wb.active

    # Update the existing cells with new data
    ws['A2'] = username


    # Save the workbook
    wb.save("TestData/LoginData.xlsx")

    # BannerPath = "C:/Users/HP/PycharmProjects/AutomationFramework/imageFiles/3.jpg"
    # Upload_AgainBanner = "C:/Users/HP/PycharmProjects/AutomationFramework/imageFiles/5.jpg"
    # ProfilePath = "C:/Users/HP/PycharmProjects/AutomationFramework/imageFiles/4.jpg"
    # UpdateProfile_Again = "C:/Users/HP/PycharmProjects/AutomationFramework/imageFiles/7.jpg"

    Companyname = randomGen.random_addressInput() #"Econtent"
    orgname = randomGen.random_first_name()    #"Software and IT"
    industry = "educational services"
    website = "https://www.google.com/"
    companysummary = randomGen.random_overviwDescription()
    address = randomGen.random_addressInput()
    pincode = randomGen.random_pinCode()
    domainname = "mailcatch.com"
    contactperson = randomGen.random_first_name()
    url = "https://www.instagram.com/"

    # awards_path = "C:/Users/HP/PycharmProjects/AutomationFramework/imageFiles/11.jpg"
    awardTitle = randomGen.random_addressInput()

    logger = LogGen.loggen()



    @pytest.mark.regression
    @pytest.mark.run(order=1)
    # @pytest.mark.skip(reason="skipping this test")
    def test_BannerImage(self):
        self.lp = LoginPage(self.driver)
        self.lp.setUserName(self.username)
        self.lp.setpassword(self.password)
        self.lp.clickLogin()
        self.lp.clickCompanyProfile()
        self.logger.info("****** TC_01	Verify the Banner image/upload/save/edit/delete *****")
        self.lp.BannerImageClick(self.absolute_path3)
        self.lp.SaveBannerImage()
        # time.sleep(3)

        xpath = "// div[contains(text(), 'Banner image uploaded successfully.')]"
        try:
            # Use WebDriverWait to wait for the element to be present
            element = WebDriverWait(self.driver, 10).until(
                EC.presence_of_element_located((By.XPATH, xpath))
            )
            self.logger.info(f"Text Found : {element.text}")
            assert True
        except:
            self.logger.info(f"Text Not Found")
            self.driver.save_screenshot(".\\ScreenShots\\" + "test_MediaDriveVerify.png")
            assert False

        self.my = MyprofilePage(self.driver)
        time.sleep(1)
        self.my.clickClosetoaster()
        # time.sleep(3)
        self.lp.EditBanner()
        self.lp.UploadBannerImageAgain(self.absolute_path4)
        self.lp.SaveBannerImage()
        xpath = "// div[contains(text(), 'Banner image uploaded successfully.')]"
        try:
            # Use WebDriverWait to wait for the element to be present
            element = WebDriverWait(self.driver, 10).until(
                EC.presence_of_element_located((By.XPATH, xpath))
            )
            self.logger.info(f"Text Found : {element.text}")
            assert True
        except:
            self.logger.info(f"Text Not Found")
            self.driver.save_screenshot(".\\ScreenShots\\" + "test_MediaDriveVerify.png")
            assert False

        # time.sleep(3)
        time.sleep(1)
        self.my.clickClosetoaster()
        self.lp.EditBanner()
        self.lp.BannerImageRemove()
        # time.sleep(3)
        xpath = "// div[contains(text(), 'Banner image removed successfully')]"
        try:
            # Use WebDriverWait to wait for the element to be present
            element = WebDriverWait(self.driver, 10).until(
                EC.presence_of_element_located((By.XPATH, xpath))
            )
            self.logger.info(f"Text Found : {element.text}")
            assert True
        except:
            self.logger.info(f"Text Not Found")
            self.driver.save_screenshot(".\\ScreenShots\\" + "test_MediaDriveVerify.png")
            assert False

    @pytest.mark.run(order=2)
    # @pytest.mark.test
    @pytest.mark.regression
    # @pytest.mark.skip(reason="skipping this test")
    def test_ProfileImage(self):
        self.lp = LoginPage(self.driver)
        self.lp.setUserName(self.username)
        self.lp.setpassword(self.password)
        self.lp.clickLogin()
        self.lp.clickCompanyProfile()
        self.logger.info("****** TC_03	Verify the Profile  image Upload/edit/save/delete  *****")
        self.lp.ProfileImageClick(self.absolute_path5)
        self.lp.SaveProfileImage()

        xpath = "// div[contains(text(), 'Profile image uploaded successfully.')]"
        try:
            # Use WebDriverWait to wait for the element to be present
            element = WebDriverWait(self.driver, 10).until(
                EC.presence_of_element_located((By.XPATH, xpath))
            )
            self.logger.info(f"Text Found : {element.text}")
            assert True
        except:
            self.logger.info(f"Text Not Found")
            self.driver.save_screenshot(".\\ScreenShots\\" + "test_MediaDriveVerify.png")
            assert False

        self.lp.EditProfile()
        self.lp.UploadProfileImage(self.absolute_path6)
        self.lp.SaveProfileImage()

        xpath = "//div[contains(text(),'Profile image uploaded successfully.')]"
        try:
            # Use WebDriverWait to wait for the element to be present
            element = WebDriverWait(self.driver, 10).until(
                EC.presence_of_element_located((By.XPATH, xpath))
            )
            self.logger.info(f"Text Found : {element.text}")
            assert True
        except:
            self.logger.info(f"Text Not Found")
            self.driver.save_screenshot(".\\ScreenShots\\" + "test_MediaDriveVerify.png")
            assert False
        # self.my = MyprofilePage(self.driver)
        # time.sleep(1)
        # self.my.clickClosetoaster()
        self.lp.EditProfile()
        self.lp.ProfileImageRemove()
        # time.sleep(4)
        xpath = "// div[contains(text(), 'Profile image removed successfully.')]"
        try:
            # Use WebDriverWait to wait for the element to be present
            element = WebDriverWait(self.driver, 10).until(
                EC.presence_of_element_located((By.XPATH, xpath))
            )
            self.logger.info(f"Text Found : {element.text}")
            assert True
        except:
            self.logger.info(f"Text Not Found")
            self.driver.save_screenshot(".\\ScreenShots\\" + "test_MediaDriveVerify.png")
            assert False

    @pytest.mark.run(order=3)
    @pytest.mark.regression
    # @pytest.mark.skip(reason="skipping this test")
    def test_OfficialDetails(self):
        self.lp = LoginPage(self.driver)
        self.lp.setUserName(self.username)
        self.lp.setpassword(self.password)
        self.lp.clickLogin()
        self.lp.clickCompanyProfile()
        self.logger.info("****** TC_05	Verify the displaying all the Official details  *****")
        self.lp.ofc_details()
        # self.lp.setCompany(self.Companyname)
        self.lp.setOrganisation(self.orgname)
        self.lp.RemoveIndustry()
        self.lp.setIndustry(self.industry)
        self.lp.clickIndustryName()
        self.lp.clickCalendar()
        self.lp.clickonpreviousmonth()
        self.lp.clickDate()
        self.lp.setwebsite(self.website)
        self.lp.clicksave()

        xpath = "// div[contains(text(), 'Profile updated  successfully')]"
        try:
            # Use WebDriverWait to wait for the element to be present
            element = WebDriverWait(self.driver, 10).until(
                EC.presence_of_element_located((By.XPATH, xpath))
            )
            self.logger.info(f"Text Found : {element.text}")
            assert True
        except:
            self.logger.info(f"Text Not Found")
            self.driver.save_screenshot(".\\ScreenShots\\" + "test_MediaDriveVerify.png")
            assert False

    @pytest.mark.run(order=4)
    # @pytest.mark.test
    @pytest.mark.regression
    # @pytest.mark.skip(reason="skipping this test")
    def test_OverView(self):
        self.lp = LoginPage(self.driver)
        self.lp.setUserName(self.username)
        self.lp.setpassword(self.password)
        self.lp.clickLogin()
        # self.lp.clickNewsfeedModule()
        self.lp.clickCompanyProfile()
        self.lp.scrollIntoOverView()

        self.logger.info("****** TC_11	Verify the Overview, by Add, Save and Update  *****")
        time.sleep(2)
        self.lp.clickEdit()
        self.lp.setCompanySummary(self.companysummary)
        self.lp.setAddress(self.address)
        actions = ActionChains(self.driver)
        actions.send_keys(Keys.PAGE_DOWN).perform()
        self.driver.execute_script("window.scrollBy(0, 500);")
        time.sleep(2)
        self.lp.countrydropdown()
        self.lp.clickonselectindia()
        self.lp.clickState()
        self.lp.clickstatelistbox()
        self.lp.selectState()
        self.lp.clickcity()
        self.lp.selectCity()
        self.lp.setpincode(self.pincode)
        self.lp.setdomainName(self.domainname)
        self.lp.setcontactPerson(self.contactperson)
        self.lp.clicksave()
        xpath = "// div[contains(text(), 'Company profile updated')]"
        try:
            # Use WebDriverWait to wait for the element to be present
            element = WebDriverWait(self.driver, 10).until(
                EC.presence_of_element_located((By.XPATH, xpath))
            )
            self.logger.info(f"Text Found : {element.text}")
            assert True
        except:
            self.logger.info(f"Text Not Found")
            self.driver.save_screenshot(".\\ScreenShots\\" + "test_MediaDriveVerify.png")
            assert False


    @pytest.mark.run(order=5)
    # @pytest.mark.test
    @pytest.mark.sunitha
    # @pytest.mark.skip(reason="skipping this test")
    def test_Awards(self):
        self.lp = LoginPage(self.driver)
        self.lp.setUserName(self.username)
        self.lp.setpassword(self.password)
        self.lp.clickLogin()
        self.lp.clickCompanyProfile()
        self.lp.scrollIntoAwards()
        self.logger.info("****** TC_13	Verify the Awards, by Add, Save and Edit *****")
        self.lp.AwardsEdit()
        self.lp.ClickPreview(self.absolute_path7)
        self.lp.TitleInput(self.awardTitle)
        self.lp.SaveAward()
        xpath = "// div[contains(text(), 'Award added successfully')]"
        try:
            # Use WebDriverWait to wait for the element to be present
            element = WebDriverWait(self.driver, 10).until(
                EC.presence_of_element_located((By.XPATH, xpath))
            )
            self.logger.info(f"Text Found : {element.text}")
            assert True
        except:
            self.logger.info(f"Text Not Found")
            self.driver.save_screenshot(".\\ScreenShots\\" + "test_MediaDriveVerify.png")
            assert False

    @pytest.mark.run(order=6)
    @pytest.mark.flaky(rerun=3, rerun_delay=2)
    @pytest.mark.regression
    # @pytest.mark.skip(reason="skipping this test")
    def test_socialMediaLinks(self):
        self.lp = LoginPage(self.driver)
        self.lp.setUserName(self.username)
        self.lp.setpassword(self.password)
        self.lp.clickLogin()
        self.lp.clickCompanyProfile()
        self.logger.info("****** TC_16	Verify Social Media Links details, by Add, Save and Update  *****")
        self.lp.clicksocialMediaLinks()
        self.lp.clicksocilmedialist()
        self.lp.clickSocialMediaName()
        self.lp.setUrl(self.url)
        self.lp.clickUrlsave()
        xpath = "// div[contains(text(), 'Profile updated  successfully')]"
        try:
            # Use WebDriverWait to wait for the element to be present
            element = WebDriverWait(self.driver, 10).until(
                EC.presence_of_element_located((By.XPATH, xpath))
            )
            self.logger.info(f"Text Found : {element.text}")
            assert True
        except:
            self.logger.info(f"Text Not Found")
            self.driver.save_screenshot(".\\ScreenShots\\" + "test_MediaDriveVerify.png")
            assert False


    @pytest.mark.run(order=7)
    @pytest.mark.falky(rerun=3, rerun_delay=2)
    @pytest.mark.regression
    def test_ClickingElements(self):
        self.lp = LoginPage(self.driver)
        # time.sleep(2)
        self.lp.setUserName(self.username)
        self.lp.setpassword(self.password)
        # time.sleep(2)
        self.lp.clickLogin()
        # time.sleep(2)
        #self.lp.clickNewsfeedModule()
        # time.sleep(2)
        self.lp.clickCompanyProfile()

        time.sleep(1)
        self.logger.info("****** TC_19	Verify Networks by click on that  *****")
        self.lp.NetworksClick()
        xpath = "//button[normalize-space()='MY NETWORKS']"
        try:
            # Use WebDriverWait to wait for the element to be present
            element = WebDriverWait(self.driver, 10).until(
                EC.presence_of_element_located((By.XPATH, xpath))
            )
            self.logger.info(f"Text Found : {element.text}")
            assert True
        except:
            self.logger.info(f"Text Not Found")
            self.driver.save_screenshot(".\\ScreenShots\\" + "test_MediaDriveVerify.png")
            assert False

        # time.sleep(2)
        self.driver.back()
        # time.sleep(3)
        self.logger.info("****** TC_20	verify the certificate by click on that  *****")
        self.lp.CertificationsClick()
        xpath = "//button[normalize-space()='Create certification']"
        try:
            # Use WebDriverWait to wait for the element to be present
            element = WebDriverWait(self.driver, 10).until(
                EC.presence_of_element_located((By.XPATH, xpath))
            )
            self.logger.info(f"Text Found : {element.text}")
            assert True
        except:
            self.logger.info(f"Text Not Found")
            self.driver.save_screenshot(".\\ScreenShots\\" + "test_MediaDriveVerify.png")
            assert False

        self.driver.back()
        time.sleep(3)
        self.logger.info("****** TC_18	Verify Resources by click on that *****")
        self.lp.ResourceClick()
        xpath = "//button[normalize-space()='My RESOURCES']"
        try:
            # Use WebDriverWait to wait for the element to be present
            element = WebDriverWait(self.driver, 10).until(
                EC.presence_of_element_located((By.XPATH, xpath))
            )
            self.logger.info(f"Text Found : {element.text}")
            assert True
        except:
            self.logger.info(f"Text Not Found")
            self.driver.save_screenshot(".\\ScreenShots\\" + "test_MediaDriveVerify.png")
            assert False

        time.sleep(2)
        self.driver.back()
        # self.driver.back()
        self.logger.info("****** TC_25	verify the setting tab by clicking on it *****")
        self.lp.Settings()
        self.logger.info("****** TC_24	verify the news feed by clicking on that  *****")
        self.lp.NewsFeed()


