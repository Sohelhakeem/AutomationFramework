import os
import time
import pytest
from openpyxl.reader.excel import load_workbook

from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC

from sunithaPageObjects.MyProfile import MyprofilePage
from sunithaPageObjects.CompanyProfile import LoginPage
from pageObjects.randomGen import randomGen
from utilities.customLogger import LogGen
from utilities.readProperties import ReadConfig
from pageObjects.EmployeeModulePage import EmployeeModulePage
from GenericLib.BaseClass import BaseClass

class TestMyProfile(BaseClass):
    baseURL = ReadConfig.getApplicationURL()
    # username = ReadConfig.getUseremail()
    password = ReadConfig.getPassword()
    relative_three = "Files/three.png"
    absolute_path3 = os.path.abspath(relative_three)
    relative_four = "Files/four.png"
    absolute_path4 = os.path.abspath(relative_four)
    relative_five = "Files/five.png"
    absolute_path5 = os.path.abspath(relative_five)
    relative_six = "Files/six.jpg"
    absolute_path6 = os.path.abspath(relative_six)
    overviewTextarea = "InLinks is the first"
    firstName = randomGen.random_first_name()
    # firstName = "Sunitha"
    lastName = randomGen.random_last_name()
    # lastName = "K"
    emailAddress = randomGen.random_email()
    # emailAddress = "sunitha@peoplelinkvc.com"
    phone = randomGen.random_phone_number()
    phoneNumber = phone
    # phoneNumber = phone+"2"
    # degree = randomGen.random_degree()  # "B.Sc"
    degree = "B.Sc"
    specialization = randomGen.random_specialization()
    # specialization = "Computer science"
    university = randomGen.random_university()
    # university = "Acharya Nagarguna Univercity"
    savEbutton = "//button[normalize-space()='Save']"
    addressInput = randomGen.random_addressInput()
    # addressInput = "Hyderabad,Hitech city,Cyber towers"
    pinCode = randomGen.random_pinCode()
    urlInput = "www.instagram.com"
    overviewText = randomGen.random_overviwDescription()
    # overviewTest = "PeopleLink provides solutions for various room types, including personal, huddle, conference, training & board that facilitate local or remote meetings using high-quality AV solutions, plug & play, remote control, and instant content sharing."
    empid = "8679"
    workbook = load_workbook("TestData/LoginData.xlsx")
    # Access the active worksheet
    worksheet = workbook.active
    username = worksheet["A2"].value
    usernames4 = worksheet["E2"].value
    usernames5 = worksheet["I2"].value
    workbook.close()
    # Load the existing workbook
    wb = load_workbook("TestData/LoginData.xlsx")
    # Select the active worksheet
    ws = wb.active
    # Update the existing cells with new data
    ws['A2'] = username
    ws['E2'] = usernames4
    ws['I2'] = usernames5

    # Save the workbook
    wb.save("TestData/LoginData.xlsx")

    logger = LogGen.loggen()

    @pytest.mark.run(order=28)
    @pytest.mark.regression
    @pytest.mark.flaky(rerun=3, rerun_delay=2)
    # @pytest.mark.skip(reason="skipping this test")
    def test_BannerImage(self):  # def test_loginTitle(self):self #def test_ProfileUploading
        self.lp = LoginPage(self.driver)
        self.lp.setUserName(self.username)
        self.lp.setpassword(self.password)
        self.lp.clickLogin()
        self.my = MyprofilePage(self.driver)
        self.my.clickMyProfileModule()

        # Banner image uploading----------------------------------------------------------------------
        self.logger.info("****** TC_24	Verify the Profile  Banner image upload/update/save/delete  *****")
        self.my.uploadBannerImage(self.absolute_path3)
        self.my.SaveBannerImage()

        xpath = "//div[contains(text(),'Banner image uploaded successfully.')]"
        try:
            # Use WebDriverWait to wait for the element to be present
            element = WebDriverWait(self.driver, 10).until(
                EC.presence_of_element_located((By.XPATH, xpath))
            )
            self.logger.info(f"Text Found : {element.text}")
            assert True
        except:
            self.logger.info(f"Text Not Found")
            self.driver.save_screenshot(".\\ScreenShots\\" + "test_MediaDriveVerify.png")
            assert False

        time.sleep(1)
        self.my.clickClosetoaster()
        self.my.BannerImageEdit()
        self.my.BannerImageUpdate(self.absolute_path4)
        self.my.SaveBannerImage()
        # time.sleep(1)

        xpath = "// div[contains(text(), 'Banner image uploaded successfully.')]"
        try:
            # Use WebDriverWait to wait for the element to be present
            element = WebDriverWait(self.driver, 10).until(
                EC.presence_of_element_located((By.XPATH, xpath))
            )
            self.logger.info(f"Text Found : {element.text}")
            assert True
        except:
            self.logger.info(f"Text Not Found")
            self.driver.save_screenshot(".\\ScreenShots\\" + "test_MediaDriveVerify.png")
            assert False
        time.sleep(1)
        self.my.clickClosetoaster()
        self.my.BannerImageEdit()
        self.my.BannerImageRemove()
        xpath = "//div[contains(text(),'Banner image removed successfully')]"
        try:
            # Use WebDriverWait to wait for the element to be present
            element = WebDriverWait(self.driver, 10).until(
                EC.presence_of_element_located((By.XPATH, xpath))
            )
            self.logger.info(f"Text Found : {element.text}")
            assert True
        except:
            self.logger.info(f"Text Not Found")
            self.driver.save_screenshot(".\\ScreenShots\\" + "test_MediaDriveVerify.png")
            assert False




        # Profile image uploading-----------------------------------------------------------------------

    @pytest.mark.run(order=29)
    @pytest.mark.regression
    @pytest.mark.krishna
    # @pytest.mark.skip(reason="skipping this test")
    def test_profileImages(self):
        self.lp = LoginPage(self.driver)
        self.lp.setUserName(self.username)
        self.lp.setpassword(self.password)
        self.lp.clickLogin()
        self.my = MyprofilePage(self.driver)
        self.my.clickMyProfileModule()
        time.sleep(3)
        self.logger.info("****** TC_24	Verify the Profile  Banner image upload/update/save/delete  *****")
        self.my.setprofileImage(self.absolute_path5)

        self.my.saveProfileImage()
        time.sleep(1)
        xpath = "//div[contains(text(),'Profile image uploaded successfully.')]"
        try:
            # Use WebDriverWait to wait for the element to be present
            element = WebDriverWait(self.driver, 10).until(
                EC.presence_of_element_located((By.XPATH, xpath))
            )
            self.logger.info(f"Text Found : {element.text}")
            assert True
        except:
            self.logger.info(f"Text Not Found")
            self.driver.save_screenshot(".\\ScreenShots\\" + "test_MediaDriveVerify.png")
            assert False
        time.sleep(1)
        self.my.ProfileEditButton()
        self.my.ProfileUpdate(self.absolute_path6)
        self.my.saveProfileImage()
        time.sleep(1)
        xpath = "//div[contains(text(),'Profile image uploaded successfully.')]"
        try:
            # Use WebDriverWait to wait for the element to be present
            element = WebDriverWait(self.driver, 10).until(
                EC.presence_of_element_located((By.XPATH, xpath))
            )
            self.logger.info(f"Text Found : {element.text}")
            assert True
        except:
            self.logger.info(f"Text Not Found")
            self.driver.save_screenshot(".\\ScreenShots\\" + "test_MediaDriveVerify.png")
            assert False
        time.sleep(1)
        self.my.ProfileEditButton()
        self.my.ProfileRemove()
        # time.sleep(1)
        xpath = "//div[contains(text(),'Profile image removed successfully.')]"
        try:
            # Use WebDriverWait to wait for the element to be present
            element = WebDriverWait(self.driver, 10).until(
                EC.presence_of_element_located((By.XPATH, xpath))
            )
            self.logger.info(f"Text Found : {element.text}")
            assert True
        except:
            self.logger.info(f"Text Not Found")
            self.driver.save_screenshot(".\\ScreenShots\\" + "test_MediaDriveVerify.png")
            assert False





    # Required Details------------------------------------------------------------------------------
    @pytest.mark.run(order=30)
    @pytest.mark.regression
    # @pytest.mark.skip(reason="skipping this test")
    def test_RequiredDetails(self):

        self.lp = LoginPage(self.driver)
        self.lp.setUserName(self.username)
        self.lp.setpassword(self.password)
        self.lp.clickLogin()

        self.my = MyprofilePage(self.driver)
        self.my.clickMyProfileModule()
        self.logger.info("****** TC_29	Verify the displaying Required details by Edit  *****")
        self.my.clickfirstEdit()
        self.my.empID(self.empid)
        self.em = EmployeeModulePage(self.driver)
        self.em.ClickDepartmentDD()
        self.em.ClickSelectDD()
        self.em.ClickDivisionDD()
        self.em.ClickSelectDD()
        self.em.ClickDesignationDD()
        self.em.ClickSelectDD()
        self.my.updateEdit()
        xpath = "//div[contains(text(),'Profile updated  successfully')]"
        try:
            # Use WebDriverWait to wait for the element to be present
            element = WebDriverWait(self.driver, 10).until(
                EC.presence_of_element_located((By.XPATH, xpath))
            )
            self.logger.info(f"Text Found : {element.text}")
            assert True
        except:
            self.logger.info(f"Text Not Found")
            self.driver.save_screenshot(".\\ScreenShots\\" + "test_MediaDriveVerify.png")
            assert False

    # OverView___________________________________________________________________________________
    @pytest.mark.run(order=31)
    @pytest.mark.regression
    # @pytest.mark.regression
    # @pytest.mark.skip(reason="skipping this test")
    def test_OverView(self):
        self.lp = LoginPage(self.driver)
        self.lp.setUserName(self.username)
        self.lp.setpassword(self.password)
        self.lp.clickLogin()

        self.my = MyprofilePage(self.driver)
        self.my.clickMyProfileModule()
        self.logger.info("****** TC_30	Verify the Overview, by Add, Save and Update *****")
        self.my.overViewEdit()
        self.my.setoverviewTextarea(self.overviewText)
        self.my.clickOverViewSave()
        xpath = "//div[contains(text(),'Profile updated  successfully')]"
        try:
            # Use WebDriverWait to wait for the element to be present
            element = WebDriverWait(self.driver, 10).until(
                EC.presence_of_element_located((By.XPATH, xpath))
            )
            self.logger.info(f"Text Found : {element.text}")
            assert True
        except:
            self.logger.info(f"Text Not Found")
            self.driver.save_screenshot(".\\ScreenShots\\" + "test_MediaDriveVerify.png")
            assert False



        # Personal Details____________________________________________________________________________

    @pytest.mark.run(order=32)
    @pytest.mark.regression
    # @pytest.mark.krishna
    # @pytest.mark.skip(reason="skipping this test")
    def test_PersonalDetails(self):
        self.lp = LoginPage(self.driver)
        self.lp.setUserName(self.username)
        self.lp.setpassword(self.password)
        self.lp.clickLogin()

        self.my = MyprofilePage(self.driver)
        self.my.clickMyProfileModule()
        self.my.clickToScroll()
        self.logger.info("****** TC_34	Verify Personal details, by Add, Save and Update the data *****")
        self.my.clickpersonalDetails()
        self.my.setfirstName(self.firstName)
        self.my.setlastName(self.lastName)
        self.my.setEmailAddress(self.emailAddress)
        self.my.phoneNumber(self.phoneNumber)
        self.my.genderRadiobutton()
        ##### For new account we need to uncomment the below script#####
        self.my.maritalStatus()
        # time.sleep(2)
        self.my.bloodGroupClick()
        self.my.selectBloodGroup()
        self.my.saveButton()
        xpath = "//div[contains(text(),'Profile updated  successfully')]"
        try:
            # Use WebDriverWait to wait for the element to be present
            element = WebDriverWait(self.driver, 10).until(
                EC.presence_of_element_located((By.XPATH, xpath))
            )
            self.logger.info(f"Text Found : {element.text}")
            assert True
        except:
            self.logger.info(f"Text Not Found")
            self.driver.save_screenshot(".\\ScreenShots\\" + "test_MediaDriveVerify.png")
            assert False


    # Educational Details_________________________________________________________________________________
    @pytest.mark.run(order=33)
    @pytest.mark.regression
    # @pytest.mark.krishna
    # @pytest.mark.skip(reason="skipping this test")
    def test_EducationalDetails(self):
        self.lp = LoginPage(self.driver)
        self.lp.setUserName(self.username)
        self.lp.setpassword(self.password)
        self.lp.clickLogin()

        self.my = MyprofilePage(self.driver)
        self.my.clickMyProfileModule()
        self.my.clickToScroll()
        self.my.clickToScrolll()
        self.logger.info("****** TC_36	Verify Educational details, by Add, Save and Update *****")
        self.my.educatinalDetails()

        self.my.degreeField(self.degree)
        self.my.specializationField(self.specialization)
        self.my.UniversityField(self.university)
        self.my.savEbutton()
        time.sleep(3)

        xpath = "//div[contains(text(),'Profile updated  successfully')]"
        try:
            # Use WebDriverWait to wait for the element to be present
            element = WebDriverWait(self.driver, 10).until(
                EC.presence_of_element_located((By.XPATH, xpath))
            )
            self.logger.info(f"Text Found : {element.text}")
            assert True
        except:
            self.logger.info(f"Text Not Found")
            self.driver.save_screenshot(".\\ScreenShots\\" + "test_MediaDriveVerify.png")
            assert False

        # Address__________________________________________________________________________

    @pytest.mark.run(order=34)
    @pytest.mark.regression
    # @pytest.mark.krishna
    @pytest.mark.flaky(rerun=3, rerun_delay=2)
    def test_Address(self):
        self.lp = LoginPage(self.driver)
        self.lp.setUserName(self.username)
        self.lp.setpassword(self.password)
        self.lp.clickLogin()
        # self.lp.clickNewsFeedModule()

        self.my = MyprofilePage(self.driver)
        self.my.clickMyProfileModule()
        # time.sleep(3)

        self.my.clickToScrolll()
        self.my.scroll_to_end_of_page()
        # self.my.educatinalDetails()
        self.logger.info("****** TC_40	Verify Address details, by Add, Save and Update *****")
        self.my.AddressEdit()
        # time.sleep(2)
        self.my.addressField(self.addressInput)
        # time.sleep(2)
        self.my.countryField()
        # time.sleep(2)
        self.my.countryListindia()
        # time.sleep(2)
        # self.my.countryselect()

        self.my.statelistbox()
        self.my.stateSelection()
        self.my.cityListbox()
        self.my.citySelection()
        self.my.pincodeInput(self.pinCode)
        self.my.checkbox()
        time.sleep(2)
        self.my.SaveButton()
        xpath = "//div[contains(text(),'Profile updated  successfully')]"
        try:
            # Use WebDriverWait to wait for the element to be present
            element = WebDriverWait(self.driver, 10).until(
                EC.presence_of_element_located((By.XPATH, xpath))
            )
            self.logger.info(f"Text Found : {element.text}")
            assert True
        except:
            self.logger.info(f"Text Not Found")
            self.driver.save_screenshot(".\\ScreenShots\\" + "test_MediaDriveVerify.png")
            assert False

        # Social Media Links________________________________________

    @pytest.mark.run(order=35)
    @pytest.mark.regression
    # @pytest.mark.skip(reason="skipping this test")
    def test_SocialMediaLinks(self):
        self.lp = LoginPage(self.driver)
        self.lp.setUserName(self.username)
        self.lp.setpassword(self.password)
        self.lp.clickLogin()
        # time.sleep(2)
        # self.lp.clickNewsFeedModule()

        self.my = MyprofilePage(self.driver)
        self.my.clickMyProfileModule()
        self.logger.info("****** TC_38	Verify Social Media Links details, by Add, Save and Update *****")
        self.my.clickToScrolll()
        self.my.scroll_to_end_of_page()
        self.my.socialMediaLinks()
        self.my.socialMediaDropdown()
        self.my.nameSelection()
        self.my.urlField(self.urlInput)
        self.my.finalSave()
        xpath = "//div[contains(text(),'Profile updated  successfully')]"
        try:
            # Use WebDriverWait to wait for the element to be present
            element = WebDriverWait(self.driver, 10).until(
                EC.presence_of_element_located((By.XPATH, xpath))
            )
            self.logger.info(f"Text Found : {element.text}")
            assert True
        except:
            self.logger.info(f"Text Not Found")
            self.driver.save_screenshot(".\\ScreenShots\\" + "test_MediaDriveVerify.png")
            assert False
