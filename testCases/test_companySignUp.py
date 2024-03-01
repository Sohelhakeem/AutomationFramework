import time
import unittest
import pytest
from openpyxl.reader.excel import load_workbook
from openpyxl.workbook import Workbook
from selenium.common import NoSuchElementException, StaleElementReferenceException, TimeoutException
from selenium.webdriver import Keys
from selenium.webdriver.common.by import By
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from pageObjects.companyListingPage import companyListingPage
from pageObjects.LoginPage import LoginPage
from testCases.conftest import setup
from utilities.customLogger import LogGen
from utilities.readProperties import ReadConfig
from pageObjects.companySignUpPage import companySignUpPage
from pageObjects.randomGen import randomGen
from selenium import webdriver
from openpyxl import workbook
import re

class TestSignUp(unittest.TestCase):
    baseURL = ReadConfig.getApplicationURL()
    setSearchIndustryType = "Information Technology"
    password = ReadConfig.getPassword()

    def setUp(self):
        self.logger = LogGen.loggen()
        self.driver = webdriver.Chrome()  # Change to the appropriate driver
        self.driver.maximize_window()
        self.driver.implicitly_wait(20)
    def tearDown(self):
        self.driver.quit()

    @pytest.mark.run(order=1)
    @pytest.mark.regression
    @pytest.mark.test
    @pytest.mark.flaky(rerun=3, rerun_delay=2)
    def test_SignUpwithValid(self):
        self.driver.get(self.baseURL)
        self.logger.info("********TS_1	TC1_1	Verify the Signup functionality. with positive data. ***********")
        self.logger.info("******** User is on Login page ***********")

        email = randomGen.random_email()
        first_name = randomGen.random_first_name()
        company_name = randomGen.random_company_name()
        phone_number = randomGen.random_phone_number()

        self.logger.info("******** Generating and storing data into excel sheet ***********")
        # Load the existing workbook
        wb = load_workbook("TestData/LoginData.xlsx")

        # Select the active worksheet
        ws = wb.active

        # Update the existing cells with new data
        ws['A2'] = email
        ws['B2'] = first_name
        ws['C2'] = company_name
        ws['D2'] = phone_number

        # Save the workbook
        wb.save("TestData/LoginData.xlsx")

        self.sp = companySignUpPage(self.driver)
        self.sp.clicksignuplink()
        self.sp.clickCompanysignupButton()
        self.logger.info("******** user is in company signup page ***********")
        self.logger.info("******** Entering valid data into the fields ***********")
        self.sp.setCompanyName(company_name)

        self.sp.setSearchIndustryType(self.setSearchIndustryType)
        self.sp.selectCompany()
        self.sp.setContactName(first_name)
        self.sp.setEmail(email)
        self.sp.clickcountrydd()
        self.sp.clickindia()
        self.sp.clickstatedd()
        self.sp.clickTelangana()
        self.sp.clickcitydd()
        self.sp.clickHyderabad()
        self.sp.setPhone(phone_number)
        self.sp.setPassword(self.password)
        self.sp.setConfirmPassword(self.password)
        self.sp.clicktermsConditions()
        time.sleep(2)
        self.logger.info("******** Clicking on signup button ***********")
        self.logger.info("******** TC1_2  Verify that a User can Successfully Sign Up with OTP  ***********")
        self.sp.clicksignupNow()
        time.sleep(2)

        # Execute JavaScript to open a new tab
        self.driver.execute_script("window.open('about:blank', '_blank');")

        # Perform actions in the new tab (if needed)
        # For example:
        self.driver.switch_to.window(self.driver.window_handles[1])
        self.logger.info("******** Opening new url in another tab for Email OTP ***********")
        time.sleep(1)
        self.driver.get("http://mailcatch.com/en/disposable-email")
        time.sleep(1)
        yopmail = self.driver.find_element(By.XPATH, "//input[@name='box']")
        yopmail.send_keys(email + Keys.ENTER)
        time.sleep(1)

        reload_button = self.driver.find_element(By.XPATH, "//img[@title='Reload']")

        # Click the Reload button every second until the subject is displayed or a maximum time is reached
        max_wait_time = 60  # Set your maximum wait time in seconds
        start_time = time.time()

        while time.time() - start_time < max_wait_time:
            reload_button.click()

            try:
                # Check if the subject is displayed
                subject = WebDriverWait(self.driver, 10).until(
                    EC.presence_of_element_located((By.XPATH, "//td[@class='subject']"))
                )
                subject.click()
                break  # Break out of the loop if subject is displayed
            except StaleElementReferenceException:
                print("StaleElementReferenceException occurred. Retrying...")
                continue  # Retry the loop if StaleElementReferenceException occurs
            except TimeoutException:
                time.sleep(1)

        iframeElement = self.driver.find_element(By.ID, "emailframe")
        self.driver.switch_to.frame(iframeElement)

        # Code outside the loop will be executed after the loop or when a TimeoutException occurs
        otp = WebDriverWait(self.driver, 10).until(
            EC.presence_of_element_located((By.XPATH, "//body"))
        )
        self.driver.execute_script("arguments[0].scrollIntoView(true);", otp)
        time.sleep(0.5)

        confirmation_code = otp.text
        getOTP = re.search(r'\b\d+\b', confirmation_code).group()
        print(getOTP)

        # This code is for QA ENV
        otp = self.driver.find_element(By.XPATH, "//body")
        self.driver.execute_script("arguments[0].scrollIntoView(true);", otp)
        time.sleep(0.5)

        confirmation_code = otp.text
        getOTP = re.search(r'\b\d+\b', confirmation_code).group()
        print(getOTP)

        self.logger.info("******** Switching back and entering the otp ***********")
        self.driver.switch_to.default_content()

        self.driver.switch_to.window(self.driver.window_handles[0])

        self.sp.setOtp(getOTP)

        time.sleep(2)
        self.logger.info("******** TC3_1 Verify the Signup page OTP page,  Verify that a user can successfully verify their account with a valid OTP. ***********")
        self.sp.clickVerifyButton()
        self.sp.clickContinueToLogin()
        self.logger.info("******** Company Sign Up successful ***********")
        self.logger.info("******** Entering the sig up credentials for Login ***********")
        # Read data from specific cells
        email = ws['A2'].value

        self.lp = LoginPage(self.driver)
        self.lp.setUserName(email)
        self.lp.setPassword(self.password)
        self.lp.clickLogin()
        self.lp.clickcreatePost()
        self.logger.info("******** Login successful ***********")
        act_Text = self.lp.newsFeedText()

        if act_Text == "Create News Feed":
            assert True
            self.logger.info("********* SignUp Test is Passed ***********")

        else:
            self.driver.save_screenshot(".\\ScreenShots\\" + "test_SignUpwithValid.png")
            self.logger.error("********* SignUp Test is Failed ***********")
            self.driver.close()
            assert False

        time.sleep(3)
        self.driver.find_element(By.XPATH, "//div[@class='flexAutoRow alignCntr pdngHXS']").click()
        self.driver.close()

    if __name__ == '__main__':
        unittest.main(verbosity=2)


    @pytest.mark.run(order=2)
    @pytest.mark.regression
    # @pytest.mark.test
    # @pytest.mark.flaky(rerun=3, rerun_delay=2)
    def test_ListingSignUpCompany(self):
        Url = "https://preprodanalytics.inlynk.com/license"
        username = "sowjanyapreprod@yopmail.com"
        password = "Inlink@123"
        self.driver.get(Url)
        self.lp = LoginPage(self.driver)
        self.clp = companyListingPage(self.driver)
        self.clp.setUserName(username)
        self.clp.setPassword(password)
        self.lp.clickLogin()

        self.clp.clickLicense()
        wb = load_workbook("TestData/LoginData.xlsx")

        # Select the active worksheet
        ws = wb.active
        company = ws['C2'].value
        self.clp.setsearchFiled(company)
        xpath = "//span[contains(text(),'" + company + "')]"
        # Use WebDriverWait to wait for the element to be present
        element = WebDriverWait(self.driver, 10).until(
            EC.presence_of_element_located((By.XPATH, xpath))
        )

        if element:
            self.logger.info(f"Found Employee name : {element.text}")
            assert True
            # self.driver.quit()
        else:
            self.logger.info(f"Employee name not found: {element.text}")
            self.driver.save_screenshot(".\\ScreenShots\\" + "test_Employee_StatusAndRole.png")
            self.driver.close()
            self.driver.quit()
            assert False

        element.click()
        self.clp.clicksubscription()
        self.clp.clickedit()
        time.sleep(3)
        self.clp.clicklisted()
        self.clp.clickupdate()
        xpath = self.clp.ListedToast
        # Use WebDriverWait to wait for the element to be present
        element = WebDriverWait(self.driver, 10).until(
            EC.presence_of_element_located((By.XPATH, xpath))
        )

        if element:
            self.logger.info(f"Found Employee name : {element.text}")
            assert True
            # self.driver.quit()
        else:
            self.logger.info(f"Employee name not found: {element.text}")
            self.driver.save_screenshot(".\\ScreenShots\\" + "test_Employee_StatusAndRole.png")
            self.driver.close()
            self.driver.quit()
            assert False