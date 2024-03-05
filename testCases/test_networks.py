import time
import unittest

import pytest
from openpyxl.reader.excel import load_workbook
from selenium import webdriver
from selenium.common import TimeoutException
from selenium.webdriver.common.by import By
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from pageObjects.networksPage import networksPage

from pageObjects.LoginPage import LoginPage
from utilities.readProperties import ReadConfig
from utilities.customLogger import LogGen

from GenericLib.BaseClass import BaseClass

class TestNetworks(BaseClass):
    baseURL = ReadConfig.getApplicationURL()
    workbook = load_workbook("TestData/LoginData.xlsx")

    # Access the active worksheet
    worksheet = workbook.active

    username = worksheet["A2"].value
    RMname = worksheet["J2"].value
    RMname2 = worksheet["J3"].value
    RMname3 = worksheet["J4"].value
    password = ReadConfig.getPassword()
    CompanyManufacture = worksheet["H2"].value
    EmailManufacture = worksheet["I2"].value
    CompanyPartner = worksheet["H3"].value
    EmailPartner = worksheet["I3"].value
    CompanyDistributor = worksheet["H4"].value
    EmailDistributor = worksheet["I4"].value
    CompanyVendor = worksheet["H5"].value
    EmailVendor = worksheet["I5"].value

    workbook.close()

    logger = LogGen.loggen()


    @pytest.mark.run(order=1)
    @pytest.mark.regression
    @pytest.mark.skip
    @pytest.mark.flaky(reruns=3, reruns_delay=2)
    def test_RejectConnectionCompanyAsManufacturer(self):
        self.logger.info("****Started Network Connection Test****")
        wb = load_workbook("TestData/LoginData.xlsx")
        ws = wb.active
        CompEmail = ws['I2'].value
        companyName = ws['C2'].value

        self.lp = LoginPage(self.driver)
        self.logger.info(
            "Entering SuperAdmin Credentials for login username " + CompEmail + " and password " + self.password)
        self.lp.setUserName(CompEmail)
        self.lp.setPassword(self.password)

        self.lp.clickLogin()
        self.np = networksPage(self.driver)
        self.np.clickNetworks()
        self.np.setsearchField(companyName)
        time.sleep(3)
        # Verifying and Clicking on the company
        self.driver.find_element(By.XPATH, "//span[contains(text(),'" + companyName + "')]").click()
        self.np.clickConnectButton()
        time.sleep(2)
        self.np.clickDropDownList()
        self.np.clickManufacturer()
        self.np.setRM_searchField(self.RMname)
        self.np.clickSelectRM()
        self.np.clickcheckbox()
        self.np.clickButtonConnect2()
        # List of error texts to validate
        Verify_texts = [
            "Network Connections",
        ]

        # Find elements containing error texts and validate
        for text in Verify_texts:
            try:
                # Wait up to 10 seconds until at least one element with the specified text is present
                error_elements = WebDriverWait(self.driver, 10).until(
                    EC.presence_of_all_elements_located((By.XPATH, f"//*[contains(text(), '{text}')]"))
                )

                if error_elements:
                    self.logger.info(f"Found error text: {text}")
                    assert True

            except TimeoutException:
                self.logger.info(f"Error text not found within the specified time: {text}")
                self.driver.save_screenshot(".\\Screenshots\\" + "test_ConnectionCompanyAsManufacturer.png")
                assert False

        self.np.clickOKButton()
        self.lp.clickLogout()
        self.logger.info("******** Entering Requested company credentials for Network Connection ***********")
        # Read data from specific cells
        self.logger.info(
            "Entering Requested company Credentials for Approve request Username:" + CompEmail + " and Password:" + self.password)
        self.lp.setUserName(self.username)
        self.lp.setPassword(self.password)
        self.lp.clickLogin()
        self.np.clickNetworks()
        self.np.clickPendingTab()
        self.np.setsearchField(self.CompanyManufacture)
        self.driver.find_element(By.XPATH, "//span[contains(text(),'" + self.CompanyManufacture + "')]").click()
        self.np.clickRejectButton()
        self.np.setTextarea(self.CompanyManufacture)
        self.np.clickReject2Button()
        time.sleep(4)
        act_Text = self.np.Text_Connection_rejected_successfully()
        if act_Text == "Connection rejected successfully":
            assert True
            self.logger.info("********* Company Connection rejected Successfully ***********")
            # self.driver.close()
        else:
            self.driver.save_screenshot(".\\ScreenShots\\" + "connection_reject_fail.png")
            self.logger.error("********* Company reject fail ***********")
            # self.driver.close()
            assert False

    @pytest.mark.run(order=4)
    @pytest.mark.regression
    # @pytest.mark.skip(reason="skip for now")
    # @pytest.mark.flaky(reruns=3, reruns_delay=2)
    def test_ConnectionCompanyAsManufacturer(self):
        self.logger.info("****Started Network Connection Test****")
        wb = load_workbook("TestData/LoginData.xlsx")
        ws = wb.active
        CompEmail = ws['I2'].value
        companyName = ws['C2'].value
        self.lp = LoginPage(self.driver)
        self.logger.info(
            "Entering SuperAdmin Credentials for login username " + CompEmail + " and password " + self.password)
        self.lp.setUserName(CompEmail)
        self.lp.setPassword(self.password)

        self.lp.clickLogin()
        self.np = networksPage(self.driver)
        self.np.clickNetworks()
        self.np.setsearchField(companyName)
        # Verifying and Clicking on the company
        self.driver.find_element(By.XPATH, "//span[contains(text(),'" + companyName + "')]").click()
        self.np.clickConnectButton()
        time.sleep(2)
        self.np.clickDropDownList()
        self.np.clickManufacturer()
        self.np.setRM_searchField(self.RMname)
        self.np.clickSelectRM()
        self.np.clickcheckbox()
        self.np.clickButtonConnect2()
        # List of error texts to validate
        Verify_texts = [
            "Network Connections",
        ]

        # Find elements containing error texts and validate
        for text in Verify_texts:
            try:
                # Wait up to 10 seconds until at least one element with the specified text is present
                error_elements = WebDriverWait(self.driver, 10).until(
                    EC.presence_of_all_elements_located((By.XPATH, f"//*[contains(text(), '{text}')]"))
                )

                if error_elements:
                    self.logger.info(f"Found error text: {text}")
                    assert True

            except TimeoutException:
                self.logger.info(f"Error text not found within the specified time: {text}")
                self.driver.save_screenshot(".\\Screenshots\\" + "test_ConnectionCompanyAsManufacturer.png")
                assert False

        self.np.clickOKButton()
        self.lp.clickLogout()
        self.logger.info("******** Entering Requested company credentials for Network Connection ***********")
        self.logger.info(
            "Entering Requested company Credentials for Approve request Username:" + CompEmail + " and Password:" + self.password)
        self.lp.setUserName(self.username)
        self.lp.setPassword(self.password)
        self.lp.clickLogin()
        self.np.clickNetworks()
        self.np.clickPendingTab()
        self.np.setsearchField(self.CompanyManufacture)
        self.driver.find_element(By.XPATH, "//span[contains(text(),'" + self.CompanyManufacture + "')]").click()
        self.np.clickApproveButton()
        time.sleep(2)
        self.np.clickAcceptButton()
        time.sleep(2)
        Verify_texts = [
            "Network Connections",
        ]

        # Find elements containing error texts and validate
        for text in Verify_texts:
            try:
                # Wait up to 10 seconds until at least one element with the specified text is present
                error_elements = WebDriverWait(self.driver, 10).until(
                    EC.presence_of_all_elements_located((By.XPATH, f"//*[contains(text(), '{text}')]"))
                )

                if error_elements:
                    self.logger.info(f"Found error text: {text}")
                    assert True

            except TimeoutException:
                self.logger.info(f"Error text not found within the specified time: {text}")
                self.driver.save_screenshot(".\\Screenshots\\" + "test_ConnectionCompanyAsManufacturer.png")
                assert False
    @pytest.mark.run(order=5)
    @pytest.mark.regression
    # @pytest.mark.skip(reason="skip for now")
    # @pytest.mark.flaky(reruns=3, reruns_delay=2)
    def test_ConnectionCompanyAsPartner(self):
        self.logger.info("****Started Network Connection Test****")
        wb = load_workbook("TestData/LoginData.xlsx")
        ws = wb.active
        CompEmailpartner = ws['I3'].value
        companyName = ws['C2'].value
        self.lp = LoginPage(self.driver)
        self.logger.info(
            "Entering SuperAdmin Credentials for login username " + CompEmailpartner + " and password " + self.password)
        self.lp.setUserName(CompEmailpartner)
        self.lp.setPassword(self.password)

        self.lp.clickLogin()
        self.np = networksPage(self.driver)
        self.np.clickNetworks()
        self.np.setsearchField(companyName)
        # Verifying and Clicking on the company
        self.driver.find_element(By.XPATH, "//span[contains(text(),'" + companyName + "')]").click()
        self.np.clickConnectButton()
        time.sleep(2)
        self.np.clickDropDownList()
        time.sleep(2)
        self.np.clickpartner()
        self.np.setRM_searchField(self.RMname2)
        self.np.clickSelectRM()
        self.np.clickcheckbox()
        self.np.clickButtonConnect2()
        # List of error texts to validate
        Verify_texts = [
            "Network Connections",
        ]

        # Find elements containing error texts and validate
        for text in Verify_texts:
            try:
                # Wait up to 10 seconds until at least one element with the specified text is present
                error_elements = WebDriverWait(self.driver, 10).until(
                    EC.presence_of_all_elements_located((By.XPATH, f"//*[contains(text(), '{text}')]"))
                )

                if error_elements:
                    self.logger.info(f"Found error text: {text}")
                    assert True

            except TimeoutException:
                self.logger.info(f"Error text not found within the specified time: {text}")
                self.driver.save_screenshot(".\\Screenshots\\" + "test_ConnectionCompanyAsManufacturer.png")
                assert False

        self.np.clickOKButton()
        self.lp.clickLogout()
        self.logger.info("******** Entering Requested company credentials for Network Connection ***********")
        # Read data from specific cells

        self.logger.info(
            "Entering Requested company Credentials for Approve request Username:" + self.username + " and Password:" + self.password)
        self.lp.setUserName(self.username)
        self.lp.setPassword(self.password)
        self.lp.clickLogin()
        self.np.clickNetworks()
        self.np.clickPendingTab()
        self.np.setsearchField(self.CompanyPartner)
        self.driver.find_element(By.XPATH, "//span[contains(text(),'" + self.CompanyPartner + "')]").click()
        self.np.clickApproveButton()
        time.sleep(2)
        self.np.clickAcceptButton()
        time.sleep(2)
        Verify_texts = [
            "Network Connections",
        ]

        # Find elements containing error texts and validate
        for text in Verify_texts:
            try:
                # Wait up to 10 seconds until at least one element with the specified text is present
                error_elements = WebDriverWait(self.driver, 10).until(
                    EC.presence_of_all_elements_located((By.XPATH, f"//*[contains(text(), '{text}')]"))
                )

                if error_elements:
                    self.logger.info(f"Found error text: {text}")
                    assert True

            except TimeoutException:
                self.logger.info(f"Error text not found within the specified time: {text}")
                self.driver.save_screenshot(".\\Screenshots\\" + "test_ConnectionCompanyAsManufacturer.png")
                assert False

    @pytest.mark.run(order=6)
    @pytest.mark.regression
    # @pytest.mark.skip(reason="skip for now")
    # @pytest.mark.flaky(reruns=3, reruns_delay=2)
    def test_ConnectionCompanyAsShareHolder(self):
        self.logger.info("****Started Network Connection Test****")
        wb = load_workbook("TestData/LoginData.xlsx")
        ws = wb.active
        CompEmailshareholder = ws['I4'].value
        companyName = ws['C2'].value
        self.lp = LoginPage(self.driver)
        self.logger.info(
            "Entering SuperAdmin Credentials for login username " + CompEmailshareholder + " and password " + self.password)
        self.lp.setUserName(CompEmailshareholder)
        self.lp.setPassword(self.password)

        self.lp.clickLogin()
        self.np = networksPage(self.driver)
        self.np.clickNetworks()
        self.np.setsearchField(companyName)
        # Verifying and Clicking on the company
        self.driver.find_element(By.XPATH, "//span[contains(text(),'" + companyName + "')]").click()
        self.np.clickConnectButton()
        time.sleep(2)
        self.np.clickDropDownList()
        time.sleep(1)
        self.np.clickshareHolder()
        self.np.setRM_searchField(self.RMname3)
        self.np.clickSelectRM()
        self.np.clickcheckbox()
        self.np.clickButtonConnect2()
        # List of error texts to validate
        Verify_texts = [
            "Network Connections",
        ]

        # Find elements containing error texts and validate
        for text in Verify_texts:
            try:
                # Wait up to 10 seconds until at least one element with the specified text is present
                error_elements = WebDriverWait(self.driver, 10).until(
                    EC.presence_of_all_elements_located((By.XPATH, f"//*[contains(text(), '{text}')]"))
                )

                if error_elements:
                    self.logger.info(f"Found error text: {text}")
                    assert True

            except TimeoutException:
                self.logger.info(f"Error text not found within the specified time: {text}")
                self.driver.save_screenshot(".\\Screenshots\\" + "test_ConnectionCompanyAsManufacturer.png")
                assert False

        self.np.clickOKButton()
        self.lp.clickLogout()
        self.logger.info("******** Entering Requested company credentials for Network Connection ***********")
        # Read data from specific cells
        self.lp = LoginPage(self.driver)
        self.logger.info(
            "Entering Requested company Credentials for Approve request Username:" + self.EmailDistributor + " and Password:" + self.password)
        self.lp.setUserName(self.username)
        self.lp.setPassword(self.password)
        self.lp.clickLogin()
        self.np.clickNetworks()
        self.np.clickPendingTab()
        self.np.setsearchField(self.CompanyDistributor)
        self.driver.find_element(By.XPATH, "//span[contains(text(),'" + self.CompanyDistributor + "')]").click()
        self.np.clickApproveButton()
        time.sleep(4)
        self.np.clickAcceptButton()

        Verify_texts = [
            "Network Connections",
        ]

        # Find elements containing error texts and validate
        for text in Verify_texts:
            try:
                # Wait up to 10 seconds until at least one element with the specified text is present
                error_elements = WebDriverWait(self.driver, 10).until(
                    EC.presence_of_all_elements_located((By.XPATH, f"//*[contains(text(), '{text}')]"))
                )

                if error_elements:
                    self.logger.info(f"Found error text: {text}")
                    assert True

            except TimeoutException:
                self.logger.info(f"Error text not found within the specified time: {text}")
                self.driver.save_screenshot(".\\Screenshots\\" + "test_ConnectionCompanyAsManufacturer.png")
                assert False



    # @pytest.mark.smoke(order=3)
    @pytest.mark.run(order=2)
    @pytest.mark.regression
    # @pytest.mark.skip(reason="skip for now")
    # @pytest.mark.flaky(reruns=3, reruns_delay=2)
    def test_follow_and_unfollow_company(self):
        self.logger.info("****Started Network Connection Test****")
        self.logger.info("****TC_05	Verify Follow tab****")
        self.lp = LoginPage(self.driver)
        self.logger.info(
            "Entering SuperAdmin Credentials for login username " + self.username + " and password " + self.password)
        self.lp.setUserName(self.username)
        self.lp.setPassword(self.password)

        self.lp.clickLogin()
        self.np = networksPage(self.driver)
        self.np.clickNetworks()
        self.np.setsearchField(self.CompanyManufacture)
        # Verifying and Clicking on the company
        self.driver.find_element(By.XPATH, "//span[contains(text(),'" + self.CompanyManufacture + "')]").click()
        time.sleep(1)
        self.np.clickFollowButton()
        time.sleep(1)
        self.driver.back()
        self.np.clickFOLLOWTab()
        self.np.setsearchField(self.CompanyManufacture)

        if "Following" in self.driver.page_source:
            self.logger.info("********** Following company successfully *********")

        else:
            # Log and take a screenshot
            self.logger.error("************** following company is failed **********")
            self.driver.save_screenshot(".\\Screenshots\\" + "following_failed.png")
            assert False
        self.lp.clickLogout()
        self.logger.info("******** Entering Requested company credentials for Network Connection ***********")
        # Read data from specific cells
        wb = load_workbook("TestData/LoginData.xlsx")
        ws = wb.active
        CompEmail = ws['I2'].value
        companyName = ws['C2'].value
        CompanyManufacture=ws['H2'].value
        self.lp = LoginPage(self.driver)
        self.logger.info(
            "Entering Requested company Credentials for Approve request Username:" + CompEmail + " and Password:" + self.password)
        self.lp.setUserName(CompEmail)
        self.lp.setPassword(self.password)
        self.lp.clickLogin()
        self.np.clickNetworks()
        self.np.clickFOLLOWTab()
        time.sleep(1)
        self.np.setsearchField(companyName)
        self.driver.find_element(By.XPATH, "//span[contains(text(),'" + companyName + "')]").click()
        time.sleep(3)
        if "Follow" in self.driver.page_source:
            self.logger.info("********** Following company successfully *********")

        else:
            # Log and take a screenshot
            self.logger.error("************** following company is failed **********")
            self.driver.save_screenshot(".\\Screenshots\\" + "following_failed.png")
            assert False
        self.lp.clickLogout()
        self.logger.info(
            "Entering SuperAdmin Credentials for login username " + self.username + " and password " + self.password)
        self.lp.setUserName(self.username)
        self.lp.setPassword(self.password)

        self.lp.clickLogin()
        self.np = networksPage(self.driver)
        self.np.clickNetworks()
        self.np.clickFOLLOWTab()
        self.np.setsearchField(self.CompanyManufacture)
        self.driver.find_element(By.XPATH, "//span[contains(text(),'" + CompanyManufacture + "')]").click()
        time.sleep(2)
        if "Following" in self.driver.page_source:
            self.logger.info("********** Successfully see the page *********")

        else:
            # Log and take a screenshot
            self.logger.error("************** following company is failed **********")
            self.driver.save_screenshot(".\\Screenshots\\" + "following_failed.png")
            assert False
        time.sleep(2)
        self.np.clickFollowingButton()
        self.np.clickUnfollowButton()
        time.sleep(2)

    @pytest.mark.run(order=3)
    @pytest.mark.regression
    # @pytest.mark.skip(reason="skip for now")
    def test_block_and_unblock_the_followed_company(self):
        self.logger.info("****Started Network Connection Test****")
        self.logger.info("****TC_05	Verify Follow tab****")
        self.lp = LoginPage(self.driver)
        self.logger.info(
            "Entering SuperAdmin Credentials for login username " + self.username + " and password " + self.password)
        self.lp.setUserName(self.username)
        self.lp.setPassword(self.password)

        self.lp.clickLogin()
        self.np = networksPage(self.driver)
        self.np.clickNetworks()
        self.np.setsearchField(self.CompanyManufacture)
        # Verifying and Clicking on the company
        self.driver.find_element(By.XPATH, "//span[contains(text(),'" + self.CompanyManufacture + "')]").click()
        time.sleep(2)
        self.np.clickFollowButton()
        time.sleep(1)
        self.driver.back()
        self.np.clickFOLLOWTab()
        time.sleep(2)
        if "Following" in self.driver.page_source:
            self.logger.info("********** Following company successfully *********")

        else:
            # Log and take a screenshot
            self.logger.error("************** following company is failed **********")
            self.driver.save_screenshot(".\\Screenshots\\" + "following_failed.png")
            assert False
        self.lp.clickLogout()
        self.logger.info("******** Entering Requested company credentials for Network Connection ***********")
        # Read data from specific cells
        wb = load_workbook("TestData/LoginData.xlsx")
        ws = wb.active
        CompEmail = ws['I2'].value
        companyName = ws['C2'].value
        CompanyManufacture = ws['H2'].value
        self.lp = LoginPage(self.driver)
        self.logger.info(
            "Entering Requested company Credentials for Approve request Username:" + CompEmail + " and Password:" + self.password)
        self.lp.setUserName(CompEmail)
        self.lp.setPassword(self.password)
        self.lp.clickLogin()
        self.np.clickNetworks()
        self.np.clickFOLLOWTab()
        time.sleep(1)
        self.np.setsearchField(companyName)
        self.driver.find_element(By.XPATH, "//span[contains(text(),'" + companyName + "')]").click()
        time.sleep(1)
        self.driver.back()
        self.np.ClickBlockCompany()
        self.np.ClickConfirmBlock()
        time.sleep(3)
        if "You have blocked "+ companyName +"" in self.driver.page_source:
            self.logger.info("********** Following company successfully *********")

        else:
            # Log and take a screenshot
            self.logger.error("************** following company is failed **********")
            self.driver.save_screenshot(".\\Screenshots\\" + "following_failed.png")
            assert False

        self.np.clickBlocklistTab()
        time.sleep(1)
        self.np.setsearchField(companyName)
        self.driver.find_element(By.XPATH, "//span[contains(text(),'" + companyName + "')]").click()
        time.sleep(3)
        self.np.ClickUnblockCompany()
        self.np.ClickConfirmUnblock()
        time.sleep(3)
        if "You have unblocked "+ companyName +"" in self.driver.page_source:
            self.logger.info("********** Following company successfully *********")

        else:
            # Log and take a screenshot
            self.logger.error("************** following company is failed **********")
            self.driver.save_screenshot(".\\Screenshots\\" + "following_failed.png")
            assert False


        if __name__ == '__main__':
             unittest.main(verbosity=2)

