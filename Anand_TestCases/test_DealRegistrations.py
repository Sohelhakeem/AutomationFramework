import time

import pytest

from openpyxl.reader.excel import load_workbook

from pageObjects.LoginPage import LoginPage
from utilities.customLogger import LogGen
from utilities.readProperties import ReadConfig
from Anand_PageObjects.DealRegistration import dealregistration
from pageObjects.randomGen import randomGen
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC



class Test_Create_DealwithNetworkCompany:
    baseURL = ReadConfig.getApplicationURL()
    workbook = load_workbook("TestData/LoginData.xlsx")
    deptname = randomGen.random_first_name()
    comapanyname="j industries"
    nameofcompany="Instavc Technologies"
    address="Q3-A4,Cyber Towers,Hi tech city"
    department="software "+deptname
    departmenttwo="Hardware"+deptname
    departmentthree="Sales"+deptname
    departmentfour="Marketing"+deptname
    departmentfive="Networks"+deptname
    departmentsix="IT"+deptname

    wb = load_workbook("TestData/LoginData.xlsx")

    # Select the active worksheet
    ws = wb.active

    # Update the existing cells with new data
    ws['A16'] = department
    ws['C16'] = departmentfour



    # Save the workbook
    wb.save("TestData/LoginData.xlsx")
    contname="Anand"
    contemail="anand.n@instavc.com"
    contnumber="8363453672"
    nameofmanager="a"
    oppdetails="CPU & Laptops"
    currencydetails="INR"
    valueofdeal="92826"
    reason="Present deal items are not available"
    accountmanager="persona"
    search="j industries"
    searchtwo="mahindra"

    # Access the active worksheet
    worksheet = workbook.active

    username = worksheet["B14"].value
    username1=worksheet["B15"].value
    username2=worksheet["B16"].value
    password = ReadConfig.getPassword()

    workbook.close()

    logger=LogGen.loggen()

    # @pytest.mark.anand
    # @pytest.mark.skip(reason="skipping this Test")
    def test_deal_Create_Approve_Relation_Company(self,setup):
        self.logger.info("****TC_1 Verify the OEM Company Create The Deal and Approve by partner company****")
        self.driver = setup
        self.driver.get(self.baseURL)
        self.driver.maximize_window()
        self.lp = LoginPage(self.driver)
        self.lp.setUserName(self.username)
        self.lp.setPassword(self.password)
        self.lp.clickLogin()
        if "News Feed" in self.driver.page_source:
            self.logger.info("********** OEM Company Login successfully *********")

        else:
            # Log and take a screenshot
            self.logger.error("************** OEM Company Login failed **********")
            self.driver.save_screenshot(".\\Screenshots\\" + "OEM_login_fail.png")
            assert False

        self.deal=dealregistration(self.driver)
        self.deal.clickdealtab()
        if "Deal registration" in self.driver.page_source:
            self.logger.info("********** My Active Deals page & New Button will display for Creating new deal *********")

        else:
            # Log and take a screenshot
            self.logger.error("************** Deal Registration Page is not open successfully **********")
            self.driver.save_screenshot(".\\Screenshots\\" + "deal_registration_fail_to_open.png")
            assert False
        self.deal.clickonnewdeal()
        self.deal.networkcomapnyselect(self.comapanyname)
        self.deal.companydetails(self.nameofcompany)
        self.deal.adreesdetails(self.address)
        self.deal.countrydropdown()
        self.deal.selectcountry()
        self.deal.statedropdown()
        self.deal.selectstate()
        time.sleep(1)
        self.deal.citydropdown()
        self.deal.selectcity()
        self.deal.departmentname(self.departmenttwo)
        self.deal.contactname(self.contname)
        self.deal.contactemail(self.contemail)
        # self.deal.contactnumber(self.contnumber)
        self.deal.selectenterprise()
        self.deal.accountmanager(self.nameofmanager)
        self.deal.relationmanager()
        self.deal.enteropportunity(self.oppdetails)
        # self.deal.currency(self.currencydetails)
        self.deal.dealvalue(self.valueofdeal)
        self.deal.savedeal()
        if "Deal Opportunity Created" in self.driver.page_source:
            self.logger.info("********** Deal Created Successfully *********")

        else:
            # Log and take a screenshot
            self.logger.error("************** Deal not Created Successfully **********")
            self.driver.save_screenshot(".\\Screenshots\\" + "deal_creation_fail.png")
            assert False

        self.deal.okaybutton()
        time.sleep(1)
        self.lp.clickLogout()
        if "Login" in self.driver.page_source:
            self.logger.info("********** OEM Company Logout successfully *********")

        else:
            # Log and take a screenshot
            self.logger.error("************** OEM Company Logout failed **********")
            self.driver.save_screenshot(".\\Screenshots\\" + "OEM_logout_fail.png")
            assert False
        self.lp.setUserName(self.username1)
        self.lp.setPassword(self.password)
        self.lp.clickLogin()
        if "News Feed" in self.driver.page_source:
            self.logger.info("********** partner Company Login successfully *********")

        else:
            # Log and take a screenshot
            self.logger.error("************** partner Company Login failed **********")
            self.driver.save_screenshot(".\\Screenshots\\" + "partner_login_fail.png")
            assert False

        self.deal.clickdealtab()
        self.deal.networkdeals()
        if "Deal registration" in self.driver.page_source:
            self.logger.info("********** Network Deals page display  *********")

        else:
            # Log and take a screenshot
            self.logger.error("************** Network deals page not open **********")
            self.driver.save_screenshot(".\\Screenshots\\" + "network_deals_fail_to_open.png")
            assert False
        self.deal.pendingdeals()
        self.deal.searchmydealcompany(self.searchtwo)
        time.sleep(1)
        self.deal.dealcompany()
        if "Instavc Technologies" in self.driver.page_source:
            self.logger.info("********** Latest created deal is Displaying *********")

        else:
            # Log and take a screenshot
            self.logger.error("************** Not Displaying Latest Created Deal **********")
            self.driver.save_screenshot(".\\Screenshots\\" + "failed_to_open_selected_company_pending_deals.png")
            assert False
        self.deal.selectnewdeal()
        time.sleep(1)
        if "Deal Opportunity Details" in self.driver.page_source:
            self.logger.info("********** Deal created and opened successfully by that created Deal *********")

        else:
            # Log and take a screenshot
            self.logger.error("************** Failed to open created deal **********")
            self.driver.save_screenshot(".\\Screenshots\\" + "deal_open_fail.png")
            assert False
        self.deal.statusdropdown()
        self.deal.approved()
        self.deal.confirmtoapprove()
        time.sleep(2)
        if "Deal Approved Successfully" in self.driver.page_source:
            self.logger.info("********** test_deal_Create_Approved_Relation_Company successfully *********")

        else:
            # Log and take a screenshot
            self.logger.error("************** Approved deal failed  **********")
            self.driver.save_screenshot(".\\Screenshots\\" + "approved_deal_fail.png")
            assert False

    # @pytest.mark.skip(reason="skipping this Test")
    def test_deal_Create_Reject_Relation_Company(self,setup):
        self.logger.info("****TC_02 Verify the OEM Company Create The Deal and reject the deal by partner company****")
        self.driver = setup
        self.driver.get(self.baseURL)
        self.driver.maximize_window()
        self.lp = LoginPage(self.driver)
        self.lp.setUserName(self.username)
        self.lp.setPassword(self.password)
        self.lp.clickLogin()
        if "News Feed" in self.driver.page_source:
            self.logger.info("********** OEM Company Login successfully *********")

        else:
            # Log and take a screenshot
            self.logger.error("************** OEM Company Login failed **********")
            self.driver.save_screenshot(".\\Screenshots\\" + "OEM_login_fail.png")
            assert False

        self.deal=dealregistration(self.driver)
        self.deal.clickdealtab()
        if "Deal registration" in self.driver.page_source:
            self.logger.info("********** My Active Deals page & New Button will display for Creating new deal *********")

        else:
            # Log and take a screenshot
            self.logger.error("************** Deal Registration Page is not open successfully **********")
            self.driver.save_screenshot(".\\Screenshots\\" + "deal_registration_fail_to_open.png")
            assert False
        self.deal.clickonnewdeal()
        self.deal.networkcomapnyselect(self.comapanyname)
        self.deal.companydetails(self.nameofcompany)
        self.deal.adreesdetails(self.address)
        self.deal.countrydropdown()
        self.deal.selectcountry()
        self.deal.statedropdown()
        self.deal.selectstate()
        time.sleep(1)
        self.deal.citydropdown()
        self.deal.selectcity()
        self.deal.departmentname(self.departmentthree)
        self.deal.contactname(self.contname)
        self.deal.contactemail(self.contemail)
        # self.deal.contactnumber(self.contnumber)
        self.deal.selectenterprise()
        self.deal.accountmanager(self.nameofmanager)
        self.deal.relationmanager()
        self.deal.enteropportunity(self.oppdetails)
        # self.deal.currency(self.currencydetails)
        self.deal.dealvalue(self.valueofdeal)
        self.deal.savedeal()
        if "Deal Opportunity Created" in self.driver.page_source:
            self.logger.info("********** Deal Created Successfully *********")

        else:
            # Log and take a screenshot
            self.logger.error("************** Deal not Created Successfully **********")
            self.driver.save_screenshot(".\\Screenshots\\" + "deal_creation_fail.png")
            assert False

        self.deal.okaybutton()
        self.lp.clickLogout()
        if "Login" in self.driver.page_source:
            self.logger.info("********** OEM Company Logout successfully *********")

        else:
            # Log and take a screenshot
            self.logger.error("************** OEM Company Logout failed **********")
            self.driver.save_screenshot(".\\Screenshots\\" + "OEM_logout_fail.png")
            assert False
        self.lp.setUserName(self.username1)
        self.lp.setPassword(self.password)
        self.lp.clickLogin()
        if "News Feed" in self.driver.page_source:
            self.logger.info("********** partner Company Login successfully *********")

        else:
            # Log and take a screenshot
            self.logger.error("************** partner Company Login failed **********")
            self.driver.save_screenshot(".\\Screenshots\\" + "partner_login_fail.png")
            assert False
        self.deal.clickdealtab()
        self.deal.networkdeals()
        if "Deal registration" in self.driver.page_source:
            self.logger.info("********** Network Deals page display  *********")

        else:
            # Log and take a screenshot
            self.logger.error("************** Network deals page not open **********")
            self.driver.save_screenshot(".\\Screenshots\\" + "network_deals_fail_to_open.png")
            assert False
        self.deal.pendingdeals()
        self.deal.searchmydealcompany(self.searchtwo)
        time.sleep(1)
        self.deal.dealcompany()
        if "Instavc Technologies" in self.driver.page_source:
            self.logger.info("********** Latest created deal is Displaying *********")

        else:
            # Log and take a screenshot
            self.logger.error("************** Not Displaying Latest Created Deal **********")
            self.driver.save_screenshot(".\\Screenshots\\" + "failed_to_open_selected_company_pending_deals.png")
            assert False
        self.deal.selectnewdeal()
        time.sleep(1)
        if "Deal Opportunity Details" in self.driver.page_source:
            self.logger.info("********** Deal created and opened successfully by that created Deal *********")

        else:
            # Log and take a screenshot
            self.logger.error("************** Failed to open created deal **********")
            self.driver.save_screenshot(".\\Screenshots\\" + "deal_open_fail.png")
            assert False
        self.deal.statusdropdown()
        self.deal.clickreject()
        self.deal.reasontoreject(self.reason)
        self.deal.clickonreject()
        time.sleep(2)
        if "Deal Rejected Successfully" in self.driver.page_source:
            self.logger.info("********** test_deal_Create_rejected_Relation_Company successfully *********")

        else:
            # Log and take a screenshot
            self.logger.error("************** Rejected deal failed  **********")
            self.driver.save_screenshot(".\\Screenshots\\" + "rejected_deal_fail.png")
            assert False


    # @pytest.mark.skip(reason="skipping this Test")
    def test_deal_Create_Edit_Approve_Through_Notification(self, setup):
        self.logger.info("****TC_03 Create a Deal with OEM Company and Approve the Deal through Notification ****")
        self.driver = setup
        self.driver.get(self.baseURL)
        self.driver.maximize_window()
        self.lp = LoginPage(self.driver)
        self.lp.setUserName(self.username)
        self.lp.setPassword(self.password)
        self.lp.clickLogin()
        if "News Feed" in self.driver.page_source:
            self.logger.info("********** OEM Company Login successfully *********")

        else:
            # Log and take a screenshot
            self.logger.error("************** OEM Company Login failed **********")
            self.driver.save_screenshot(".\\Screenshots\\" + "OEM_login_fail.png")
            assert False

        self.deal = dealregistration(self.driver)
        self.deal.clickdealtab()
        if "Deal registration" in self.driver.page_source:
            self.logger.info(
                "********** My Active Deals page & New Button will display for Creating new deal *********")

        else:
            # Log and take a screenshot
            self.logger.error("************** Deal Registration Page is not open successfully **********")
            self.driver.save_screenshot(".\\Screenshots\\" + "deal_registration_fail_to_open.png")
            assert False
        self.deal.clickonnewdeal()
        self.deal.networkcomapnyselect(self.comapanyname)
        self.deal.companydetails(self.nameofcompany)
        self.deal.adreesdetails(self.address)
        self.deal.countrydropdown()
        self.deal.selectcountry()
        self.deal.statedropdown()
        self.deal.selectstate()
        time.sleep(1)
        self.deal.citydropdown()
        self.deal.selectcity()
        self.deal.departmentname(self.department)
        self.deal.contactname(self.contname)
        self.deal.contactemail(self.contemail)
        # self.deal.contactnumber(self.contnumber)
        self.deal.selectenterprise()
        self.deal.accountmanager(self.nameofmanager)
        self.deal.relationmanager()
        self.deal.enteropportunity(self.oppdetails)
        # self.deal.currency(self.currencydetails)
        self.deal.dealvalue(self.valueofdeal)
        self.deal.savedeal()
        if "Deal Opportunity Created" in self.driver.page_source:
            self.logger.info("********** Deal Created Successfully *********")

        else:
            # Log and take a screenshot
            self.logger.error("************** Deal not Created Successfully **********")
            self.driver.save_screenshot(".\\Screenshots\\" + "deal_creation_fail.png")
            assert False
        self.deal.okaybutton()
        self.lp.clickLogout()
        if "Login" in self.driver.page_source:
            self.logger.info("********** OEM Company Logout successfully *********")

        else:
            # Log and take a screenshot
            self.logger.error("************** OEM Company Logout failed **********")
            self.driver.save_screenshot(".\\Screenshots\\" + "OEM_logout_fail.png")
            assert False
        self.lp.setUserName(self.username1)
        self.lp.setPassword(self.password)
        self.lp.clickLogin()

        self.deal.clickonnotification()

        # changes excel sheet data
        wb = load_workbook("TestData/LoginData.xlsx")
        ws = wb.active
        department = ws['A16'].value
        element = self.driver.find_element(By.XPATH, "//span[text()='Anand Mahindra created a new deal with Instavc Technologies in Department "+department +"']")
        # assert element.text == first_name, f"Expected '{first_name}' but found '{element.text}'"

        if element:
            self.logger.info(f"Found Employee name : {element.text}")
            assert True
            # self.driver.quit()
        else:
            self.logger.info(f"Employee name not found: {element.text}")
            self.driver.save_screenshot(".\\ScreenShots\\" + "test_ApproveSignedUpEmployee.png")
            self.driver.close()
            assert False

        element.click()
        self.deal.clickhere()
        self.deal.dealdetailspage()
        self.deal.approveediteddeal()
        self.deal.confirmupdatedeal()
        self.deal.closeapprovedtab()
        time.sleep(1)

    # @pytest.mark.skip(reason="skipping this Test")
    def test_deal_Reject_Through_Notification(self, setup):
        self.logger.info("****Started Login Test****")
        self.driver = setup
        self.driver.get(self.baseURL)
        self.driver.maximize_window()
        self.lp = LoginPage(self.driver)
        self.lp.setUserName(self.username)
        self.lp.setPassword(self.password)
        self.lp.clickLogin()
        self.deal = dealregistration(self.driver)
        self.deal.clickdealtab()
        self.deal.clickonnewdeal()
        self.deal.networkcomapnyselect(self.comapanyname)
        self.deal.companydetails(self.nameofcompany)
        self.deal.adreesdetails(self.address)
        self.deal.countrydropdown()
        self.deal.selectcountry()
        self.deal.statedropdown()
        self.deal.selectstate()
        time.sleep(1)
        self.deal.citydropdown()
        self.deal.selectcity()
        time.sleep(2)
        self.deal.departmentname(self.departmentfour)
        time.sleep(2)
        self.deal.contactname(self.contname)
        self.deal.contactemail(self.contemail)
        # self.deal.contactnumber(self.contnumber)
        self.deal.selectenterprise()
        time.sleep(2)
        self.deal.accountmanager(self.nameofmanager)
        self.deal.relationmanager()
        self.deal.enteropportunity(self.oppdetails)
        # self.deal.currency(self.currencydetails)
        self.deal.dealvalue(self.valueofdeal)
        time.sleep(1)
        self.deal.savedeal()
        self.deal.okaybutton()
        time.sleep(2)
        self.lp.clickLogout()
        time.sleep(1)
        self.lp.setUserName(self.username1)
        self.lp.setPassword(self.password)
        self.lp.clickLogin()
        self.deal.clickonnotification()

        # changes excel sheet data
        wb = load_workbook("TestData/LoginData.xlsx")
        ws = wb.active
        departmentfour = ws['C16'].value
        element = self.driver.find_element(By.XPATH, "//span[text()='Anand Mahindra created a new deal with Instavc Technologies in Department "+departmentfour +"']")
        # assert element.text == first_name, f"Expected '{first_name}' but found '{element.text}'"

        if element:
            self.logger.info(f"Found Employee name : {element.text}")
            assert True
            # self.driver.quit()
        else:
            self.logger.info(f"Employee name not found: {element.text}")
            self.driver.save_screenshot(".\\ScreenShots\\" + "test_ApproveSignedUpEmployee.png")
            self.driver.close()
            assert False
        element.click()
        time.sleep(3)
        self.deal.clickhere()
        time.sleep(2)
        self.deal.statusdropdown()
        time.sleep(2)
        self.deal.clickreject()
        time.sleep(2)
        self.deal.reasontoreject(self.reason)
        self.deal.clickonreject()
        time.sleep(3)

    # @pytest.mark.skip(reason="skipping this Test")
    def test_deal_Create_Verify_Relation_Manager_and_Approve_Check_Relation_manager(self,setup):
        self.logger.info("****Started Login Test****")
        self.driver = setup
        self.driver.get(self.baseURL)
        self.driver.maximize_window()
        self.lp = LoginPage(self.driver)
        self.lp.setUserName(self.username)
        self.lp.setPassword(self.password)
        self.lp.clickLogin()
        self.logger.info("************* Super admin Login successfully **********")
        self.deal=dealregistration(self.driver)
        self.deal.clickdealtab()
        self.deal.clickonnewdeal()
        self.deal.networkcomapnyselect(self.comapanyname)
        self.deal.companydetails(self.nameofcompany)
        self.deal.adreesdetails(self.address)
        self.deal.countrydropdown()
        self.deal.selectcountry()
        self.deal.statedropdown()
        self.deal.selectstate()
        time.sleep(1)
        self.deal.citydropdown()
        self.deal.selectcity()
        self.deal.departmentname(self.departmentfive)
        self.deal.contactname(self.contname)
        self.deal.contactemail(self.contemail)
        # self.deal.contactnumber(self.contnumber)
        self.deal.selectenterprise()
        self.deal.accountmanager(self.nameofmanager)
        self.deal.relationmanager()
        self.deal.enteropportunity(self.oppdetails)
        # self.deal.currency(self.currencydetails)
        self.deal.dealvalue(self.valueofdeal)
        self.deal.savedeal()
        self.deal.okaybutton()
        self.lp.clickLogout()
        self.lp.setUserName(self.username2)
        self.lp.setPassword(self.password)
        self.lp.clickLogin()
        self.deal.clickdealtab()
        self.deal.networkdeals()
        self.deal.pendingdeals()
        self.deal.dealcompany()
        self.deal.selectnewdeal()
        self.lp.clickLogout()
        self.lp.setUserName(self.username1)
        self.lp.setPassword(self.password)
        self.lp.clickLogin()
        self.deal.clickdealtab()
        self.deal.networkdeals()
        self.deal.pendingdeals()
        self.deal.dealcompany()
        self.deal.selectnewdeal()
        self.deal.statusdropdown()
        self.deal.approved()
        self.deal.confirmtoapprove()
        time.sleep(1)
        self.deal.closeapprovedtab()
        self.lp.clickLogout()
        self.lp.setUserName(self.username2)
        self.lp.setPassword(self.password)
        self.lp.clickLogin()
        self.deal.clickdealtab()
        self.deal.networkdeals()
        self.deal.pendingdeals()
        self.deal.dealcompany()
        self.deal.clickonactivedeals()
        self.deal.selectnewdeal()
        time.sleep(2)

    # @pytest.mark.skip(reason="skipping this Test")
    def test_deal_Create_Verify_Relation_Manager_and_Approve_Check_Relation_manager_Verify_My_Deals(self, setup):
        self.logger.info("****Started Login Test****")
        self.driver = setup
        self.driver.get(self.baseURL)
        self.driver.maximize_window()
        self.lp = LoginPage(self.driver)
        self.lp.setUserName(self.username)
        self.lp.setPassword(self.password)
        self.lp.clickLogin()
        self.logger.info("************* Super admin Login successfully **********")
        self.deal = dealregistration(self.driver)
        self.deal.clickdealtab()
        self.deal.clickonnewdeal()
        self.deal.networkcomapnyselect(self.comapanyname)
        self.deal.companydetails(self.nameofcompany)
        self.deal.adreesdetails(self.address)
        self.deal.countrydropdown()
        self.deal.selectcountry()
        self.deal.statedropdown()
        self.deal.selectstate()
        time.sleep(1)
        self.deal.citydropdown()
        self.deal.selectcity()
        self.deal.departmentname(self.departmentsix)
        self.deal.contactname(self.contname)
        self.deal.contactemail(self.contemail)
        # self.deal.contactnumber(self.contnumber)
        self.deal.selectenterprise()
        self.deal.accountmanager(self.nameofmanager)
        self.deal.relationmanager()
        self.deal.enteropportunity(self.oppdetails)
        # self.deal.currency(self.currencydetails)
        self.deal.dealvalue(self.valueofdeal)
        self.deal.savedeal()
        self.deal.okaybutton()
        self.lp.clickLogout()
        self.lp.setUserName(self.username2)
        self.lp.setPassword(self.password)
        self.lp.clickLogin()
        self.deal.clickdealtab()
        self.deal.networkdeals()
        self.deal.pendingdeals()
        self.deal.searchmydealcompany(self.searchtwo)
        time.sleep(1)
        self.deal.dealcompany()
        self.deal.selectnewdeal()
        self.lp.clickLogout()
        self.lp.setUserName(self.username1)
        self.lp.setPassword(self.password)
        self.lp.clickLogin()
        self.deal.clickdealtab()
        self.deal.networkdeals()
        self.deal.pendingdeals()
        self.deal.searchmydealcompany(self.searchtwo)
        self.deal.dealcompany()
        self.deal.selectnewdeal()
        self.deal.statusdropdown()
        self.deal.approved()
        self.deal.confirmtoapprove()
        time.sleep(1)
        self.deal.closeapprovedtab()
        self.lp.clickLogout()
        self.lp.setUserName(self.username2)
        self.lp.setPassword(self.password)
        self.lp.clickLogin()
        self.deal.clickdealtab()
        self.deal.networkdeals()
        self.deal.pendingdeals()
        self.deal.searchmydealcompany(self.searchtwo)
        time.sleep(1)
        self.deal.dealcompany()
        self.deal.clickonactivedeals()
        self.deal.selectnewdeal()
        self.lp.clickLogout()
        self.lp.setUserName(self.username)
        self.lp.setPassword(self.password)
        self.lp.clickLogin()
        self.deal.clickdealtab()
        time.sleep(1)
        self.deal.searchmydealcompany(self.search)
        time.sleep(1)
        self.deal.mydealscompany()
        self.deal.selectnewdealtwo()
        self.deal.backdealspage()
        self.deal.mypendingdeals()
        self.deal.selectnewdealtwo()
        self.deal.backdealspage()
        self.deal.myrejectdeals()
        self.deal.selectnewdealtwo()
        self.deal.backdealspage()
        self.deal.myexpiredeals()
        time.sleep(3)


