import os

import pytest
import time

from openpyxl.reader.excel import load_workbook
from selenium.common import ElementNotInteractableException
from selenium.webdriver import ActionChains
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC

from pageObjects.LoginPage import LoginPage
from krishnapageObjects.CertificationPage import Certification
from pageObjects.randomGen import randomGen
from utilities.readProperties import ReadConfig
from utilities.customLogger import LogGen

class Test_001_Certification:
    baseURL = ReadConfig.getApplicationURL()
    # username = ReadConfig.getUseremail()
    # usernames4 = ReadConfig.getuseremails4()
    # usernames5 = ReadConfig.getuseremails5()
    password = ReadConfig.getPassword()
    enterquestion = "who  is the cm of Telangana"
    firstanswer = "KCR"
    secondanswer = "Revanth Reddy"
    templatename = "General Knowledge"
    relative_five = "Files/five.png"
    absolute_path5 = os.path.abspath(relative_five)
    programname = "Gk General Knowledge"
    categoryname = "category"
    relative_three = "Files/three.png"
    absolute_path3 = os.path.abspath(relative_three)
    signname = "krishna"
    signdesignation = "QA"
    certificationdescription = "Your certification is typically displayed as a document stating that as a professional, you've been trained, educated and are prepared to meet a specific set of criteria for your role."
    questionnumber = "1"
    marks = "80"
    reapply = "1"
    workbook = load_workbook("TestData/LoginData.xlsx")

    # Access the active worksheet
    worksheet = workbook.active

    username = worksheet["A2"].value
    usernames4 = worksheet["I2"].value
    usernames5 = worksheet["B6"].value


    workbook.close()
    # Load the existing workbook
    wb = load_workbook("TestData/LoginData.xlsx")

    # Select the active worksheet
    ws = wb.active

    # Update the existing cells with new data
    ws['A2'] = username
    ws['I2'] = usernames4
    ws['B6'] = usernames5


    # Save the workbook
    wb.save("TestData/LoginData.xlsx")

    logger = LogGen.loggen()  # Logger

    @pytest.mark.babi
    @pytest.mark.regression
    @pytest.mark.run(order=1)
    # @pytest.mark.skip(reason="Skipping this test")
    def test_certificationcreatingeditinganddeleting(self, setup):
        self.logger.info("************* Test_001_Certification **********")
        self.driver = setup
        self.driver.get(self.baseURL)
        self.driver.maximize_window()

        self.lp = LoginPage(self.driver)
        self.lp.setUserName(self.username)
        self.lp.setPassword(self.password)
        self.lp.clickLogin()
        self.logger.info("************* Login succesful **********")
        acronym1 = randomGen.random_acronym1()
        acronym2 = randomGen.random_acronym2()
        certificationname = randomGen.random_certificationname()
        self.logger.info("******** Generating and storing data into excel sheet ***********")
        # Load the existing workbook
        wb = load_workbook("TestData/LoginData.xlsx")

        # Select the active worksheet
        ws = wb.active

        # Update the existing cells with new data
        ws['A21'] = certificationname

        # Save the workbook
        wb.save("TestData/LoginData.xlsx")

        self.logger.info("******* Starting acronym Test **********")
        self.cp = Certification(self.driver)
        self.cp.clickoncertificationprogramme()
        self.cp.clickonmarkingsystem()
        self.cp.clickonmarkingsystemnew()
        self.cp.setacronym1(acronym1)
        self.cp.clickonaddanotherfield()
        self.cp.setacronym2(acronym2)

        self.cp.clickonsave()
        time.sleep(3)
        if "Acronym created and unpublished successfully" in self.driver.page_source:
            self.logger.info("********** Acronym creation test is passed *********")

        else:
            # Log and take a screenshot
            self.logger.error("************** Acronym creation test is failed **********")
            self.driver.save_screenshot(".\\Screenshots\\" + "test_acronymcreation.png")
            assert False
        time.sleep(3)
        self.cp.setsearch(acronym1)
        self.cp.clickonacronymedit()
        self.cp.clickonacronympublish()
        time.sleep(3)
        if "Acronym updated and published successfully" in self.driver.page_source:
            self.logger.info("********** Acronym update test is passed *********")

        else:
            # Log and take a screenshot
            self.logger.error("************** Acronym update test is failed **********")
            self.driver.save_screenshot(".\\Screenshots\\" + "test_acronymupdate.png")
            assert False
        time.sleep(3)
        self.cp.clickonquestionbank()
        self.cp.clickonmarkingsystemnew()
        self.cp.setenterquestion(self.enterquestion)
        # self.cp.clickonselectmarkingsystem()
        # self.cp.clickonselectmarkingsystemoption()
        # self.cp.clickonselectacronym()
        self.cp.setfirstanswer(self.firstanswer)
        self.cp.clickonadd()
        self.cp.setsecondanswer(self.secondanswer)
        self.cp.clickonselectanswer()
        self.cp.clickonselectmarkingsystem()
        self.cp.clickonselectmarkingsystemoption()
        self.cp.clickonselectacronym()
        self.cp.clickonquestionsave()
        time.sleep(3)
        if "Question created and unpublished successfully" in self.driver.page_source:
            self.logger.info("********** Question creation test is passed *********")

        else:
            # Log and take a screenshot
            self.logger.error("************** Question creation test is failed **********")
            self.driver.save_screenshot(".\\Screenshots\\" + "test_questioncreation.png")
            assert False
        time.sleep(3)
        self.cp.setsearch(self.enterquestion)
        self.cp.clickonquestionedit()
        self.cp.clickonbuttonedit()
        self.cp.clickonquestionpublish()
        time.sleep(3)
        if "Question update and published successfully" in self.driver.page_source:
            self.logger.info("********** Question update test is passed *********")

        else:
            # Log and take a screenshot
            self.logger.error("************** Question update test is failed **********")
            self.driver.save_screenshot(".\\Screenshots\\" + "test_questionupdate.png")
            assert False
        time.sleep(1)
        self.cp.clickontemplates()
        self.cp.clickonmarkingsystemnew()
        time.sleep(1)
         # Locate the element to perform mouse hover
        hover_element = self.driver.find_element(By.XPATH, "//div[@class='templatemediaOverlay certHover']")

        # Perform mouse hover action
        action = ActionChains(self.driver)
        action.move_to_element(hover_element).perform()

        # Wait for a brief moment if needed
        # You can use WebDriverWait for this if the page takes time to load after hover

        # Locate and click on the edit button
        edit_button = WebDriverWait(self.driver, 10).until(
            EC.element_to_be_clickable((By.XPATH, "/html[1]/body[1]/div[1]/div[1]/div[1]/div[2]/div[2]/div[3]/div[1]/div[2]/div[1]/div[1]/div[1]/div[1]/button[1]"))
         )
        edit_button.click()
        self.cp.settemplatename(self.templatename)
        self.cp.setlogoimg(self.absolute_path5)
        self.cp.clickonimgsave()
        self.cp.setprogramname(self.programname)
        self.cp.setcategoryname(self.categoryname)
        self.cp.setsign(self.absolute_path3)
        self.cp.clickonimgsave()
        self.cp.setsignname(self.signname)
        self.cp.setsigndesignation(self.signdesignation)
        self.cp.clickontemplatesave()
        time.sleep(3)
        if "Created and unpublished successfully" in self.driver.page_source:
            self.logger.info("********** template creation test is passed *********")

        else:
            # Log and take a screenshot
            self.logger.error("************** template creation test is failed **********")
            self.driver.save_screenshot(".\\Screenshots\\" + "test_templatecreation.png")
            assert False
        time.sleep(3)
        self.cp.clickonedittemplate()
        self.cp.clickontemplatepublish()
        time.sleep(3)
        if "Updated and published successfully" in self.driver.page_source:
            self.logger.info("********** template update test is passed *********")

        else:
            # Log and take a screenshot
            self.logger.error("************** template update test is failed **********")
            self.driver.save_screenshot(".\\Screenshots\\" + "test_templateupdate.png")
            assert False
        time.sleep(3)
        self.cp.clickoncertification()
        self.cp.clickonmarkingsystemnew()
        self.cp.setcertificationname(certificationname)
        self.cp.setcertificationdescription(self.certificationdescription)
        self.cp.clickonpublic()
        self.cp.clickonselectmarkingsystemfield()
        self.cp.clickonacronymselect()
        # time.sleep(3)
        self.cp.clickonquestionnumber()
        self.cp.setquestionnumber(self.questionnumber)
        self.cp.setMarks(self.marks)
        self.cp.setreapply(self.reapply)
        self.cp.clickonminutes()
        self.cp.clickonminutestime()
        self.cp.clickonoptions()
        self.cp.clickonselectoption()
        self.cp.clickonselecttemplate()
        self.cp.clickontemplateclick()
        self.cp.clickoncertificatesave()
        self.cp.clickoncertificateconfirmsave()
        time.sleep(3)
        if "Created and unpublished successfully" in self.driver.page_source:
            self.logger.info("********** certification creation test is passed *********")

        else:
            # Log and take a screenshot
            self.logger.error("************** certification creation test is failed **********")
            self.driver.save_screenshot(".\\Screenshots\\" + "test_certificationcreation.png")
            assert False
        time.sleep(3)
        self.cp.setsearch(certificationname)
        time.sleep(2)
        self.driver.find_element(By.XPATH, "//span[text()='" + certificationname + "']").click()
        self.cp.clickoncertificateedit()
        self.cp.clickoncertificatepublish()
        self.cp.clickoncertificateconfirmpublish()
        time.sleep(3)
        if "Updated and published successfully" in self.driver.page_source:
            self.logger.info("********** certification update test is passed *********")


        else:
            # Log and take a screenshot
            self.logger.error("************** certification update test is failed **********")
            self.driver.save_screenshot(".\\Screenshots\\" + "test_certificateupdate.png")
            assert False
        time.sleep(3)
        self.cp.setsearch(certificationname)
        time.sleep(2)
        self.driver.find_element(By.XPATH, "//span[text()='" + certificationname + "']").click()
        self.cp.clickoncertificatedelete()
        self.cp.clickoncertificateconfirmdelete()
        time.sleep(3)
        if "Deleted successfully" in self.driver.page_source:
            self.logger.info("********** certification delete test is passed *********")

        else:
            # Log and take a screenshot
            self.logger.error("************** certification delete test is failed **********")
            self.driver.save_screenshot(".\\Screenshots\\" + "test_certificatedelete.png")
            assert False
        time.sleep(3)
        self.cp.clickontemplates()
        self.cp.clickontemplatedelete()
        self.cp.clickoncertificateconfirmdelete()
        time.sleep(3)
        if "Deleted successfully" in self.driver.page_source:
            self.logger.info("********** template delete test is passed *********")


        else:
            # Log and take a screenshot
            self.logger.error("************** template delete test is failed **********")
            self.driver.save_screenshot(".\\Screenshots\\" + "test_templatedelete.png")
            assert False
        time.sleep(3)
        self.cp.clickonquestionbank()
        self.cp.clickonquestionedit()
        self.cp.clickonquestionbankdelete()
        self.cp.clickoncertificateconfirmdelete()
        time.sleep(3)
        if "Deleted successfully" in self.driver.page_source:
            self.logger.info("********** question delete test is passed *********")


        else:
            # Log and take a screenshot
            self.logger.error("************** question delete test is failed **********")
            self.driver.save_screenshot(".\\Screenshots\\" + "test_questiondelete.png")
            assert False
        time.sleep(3)
        self.cp.clickonmarkingsystem()
        self.cp.clickonacronymdelete()
        self.cp.clickoncertificateconfirmdelete()
        time.sleep(3)
        if "Acronym deleted successfully" in self.driver.page_source:
            self.logger.info("********** Acronym delete test is passed *********")
            self.driver.close()

        else:
            # Log and take a screenshot
            self.logger.error("************** Acronym delete test is failed **********")
            self.driver.save_screenshot(".\\Screenshots\\" + "test_acronymdelete.png")
            assert False
        time.sleep(3)

    @pytest.mark.sanity
    @pytest.mark.regression
    @pytest.mark.run(order=2)
    # @pytest.mark.skip(reason="Skipping this test")
    def test_certificationcreating(self, setup):
        self.logger.info("************* Test_002_Certification **********")
        self.driver = setup
        self.driver.get(self.baseURL)
        self.driver.maximize_window()

        self.lp = LoginPage(self.driver)
        self.lp.setUserName(self.username)
        self.lp.setPassword(self.password)
        self.lp.clickLogin()
        self.logger.info("************* Login succesful **********")
        acronym1 = randomGen.random_acronym1()
        acronym2 = randomGen.random_acronym2()
        certificationname = randomGen.random_certificationname()
        self.logger.info("******** Generating and storing data into excel sheet ***********")
        # Load the existing workbook
        wb = load_workbook("TestData/LoginData.xlsx")

        # Select the active worksheet
        ws = wb.active

        # Update the existing cells with new data
        ws['A21'] = certificationname

        # Save the workbook
        wb.save("TestData/LoginData.xlsx")

        self.logger.info("******* Starting acronym Test **********")
        self.cp = Certification(self.driver)
        self.cp.clickoncertificationprogramme()
        self.cp.clickonmarkingsystem()
        self.cp.clickonmarkingsystemnew()
        self.cp.setacronym1(acronym1)
        self.cp.clickonaddanotherfield()
        self.cp.setacronym2(acronym2)

        self.cp.clickonsave()
        time.sleep(3)
        if "Acronym created and unpublished successfully" in self.driver.page_source:
            self.logger.info("********** Acronym creation test is passed *********")

        else:
            # Log and take a screenshot
            self.logger.error("************** Acronym creation test is failed **********")
            self.driver.save_screenshot(".\\Screenshots\\" + "test_acronymcreation.png")
            assert False
        time.sleep(3)
        self.cp.setsearch(acronym1)
        self.cp.clickonacronymedit()
        self.cp.clickonacronympublish()
        time.sleep(3)
        if "Acronym updated and published successfully" in self.driver.page_source:
            self.logger.info("********** Acronym update test is passed *********")

        else:
            # Log and take a screenshot
            self.logger.error("************** Acronym update test is failed **********")
            self.driver.save_screenshot(".\\Screenshots\\" + "test_acronymupdate.png")
            assert False
        time.sleep(3)
        self.cp.clickonquestionbank()
        self.cp.clickonmarkingsystemnew()
        self.cp.setenterquestion(self.enterquestion)
        # self.cp.clickonselectmarkingsystem()
        # self.cp.clickonselectmarkingsystemoption()
        # self.cp.clickonselectacronym()
        self.cp.setfirstanswer(self.firstanswer)
        self.cp.clickonadd()
        self.cp.setsecondanswer(self.secondanswer)
        self.cp.clickonselectanswer()
        self.cp.clickonselectmarkingsystem()
        self.cp.clickonselectmarkingsystemoption()
        self.cp.clickonselectacronym()
        self.cp.clickonquestionsave()
        time.sleep(3)
        if "Question created and unpublished successfully" in self.driver.page_source:
            self.logger.info("********** Question creation test is passed *********")

        else:
            # Log and take a screenshot
            self.logger.error("************** Question creation test is failed **********")
            self.driver.save_screenshot(".\\Screenshots\\" + "test_questioncreation.png")
            assert False
        time.sleep(3)
        self.cp.setsearch(self.enterquestion)
        self.cp.clickonquestionedit()
        self.cp.clickonbuttonedit()
        self.cp.clickonquestionpublish()
        time.sleep(3)
        if "Question update and published successfully" in self.driver.page_source:
            self.logger.info("********** Question update test is passed *********")

        else:
            # Log and take a screenshot
            self.logger.error("************** Question update test is failed **********")
            self.driver.save_screenshot(".\\Screenshots\\" + "test_questionupdate.png")
            assert False
        time.sleep(1)
        self.cp.clickontemplates()
        self.cp.clickonmarkingsystemnew()
        time.sleep(1)
        # Locate the element to perform mouse hover
        hover_element = self.driver.find_element(By.XPATH, "//div[@class='templatemediaOverlay certHover']")

        # Perform mouse hover action
        action = ActionChains(self.driver)
        action.move_to_element(hover_element).perform()

        # Wait for a brief moment if needed
        # You can use WebDriverWait for this if the page takes time to load after hover

        # Locate and click on the edit button
        edit_button = WebDriverWait(self.driver, 10).until(
            EC.element_to_be_clickable((By.XPATH,
                                        "/html[1]/body[1]/div[1]/div[1]/div[1]/div[2]/div[2]/div[3]/div[1]/div[2]/div[1]/div[1]/div[1]/div[1]/button[1]"))
        )
        edit_button.click()
        self.cp.settemplatename(self.templatename)
        self.cp.setlogoimg(self.absolute_path5)
        self.cp.clickonimgsave()
        self.cp.setprogramname(self.programname)
        self.cp.setcategoryname(self.categoryname)
        self.cp.setsign(self.absolute_path3)
        self.cp.clickonimgsave()
        self.cp.setsignname(self.signname)
        self.cp.setsigndesignation(self.signdesignation)
        self.cp.clickontemplatesave()
        time.sleep(3)
        if "Created and unpublished successfully" in self.driver.page_source:
            self.logger.info("********** template creation test is passed *********")

        else:
            # Log and take a screenshot
            self.logger.error("************** template creation test is failed **********")
            self.driver.save_screenshot(".\\Screenshots\\" + "test_templatecreation.png")
            assert False
        time.sleep(3)
        self.cp.clickonedittemplate()
        self.cp.clickontemplatepublish()
        time.sleep(3)
        if "Updated and published successfully" in self.driver.page_source:
            self.logger.info("********** template update test is passed *********")

        else:
            # Log and take a screenshot
            self.logger.error("************** template update test is failed **********")
            self.driver.save_screenshot(".\\Screenshots\\" + "test_templateupdate.png")
            assert False
        time.sleep(3)
        self.cp.clickoncertification()
        self.cp.clickonmarkingsystemnew()
        self.cp.setcertificationname(certificationname)
        self.cp.setcertificationdescription(self.certificationdescription)
        self.cp.clickonpublic()
        self.cp.clickonselectmarkingsystemfield()
        self.cp.clickonacronymselect()
        # time.sleep(3)
        self.cp.clickonquestionnumber()
        self.cp.setquestionnumber(self.questionnumber)
        self.cp.setMarks(self.marks)
        self.cp.setreapply(self.reapply)
        self.cp.clickonminutes()
        self.cp.clickonminutestime()
        self.cp.clickonoptions()
        self.cp.clickonselectoption()
        self.cp.clickonselecttemplate()
        self.cp.clickontemplateclick()
        self.cp.clickoncertificatesave()
        self.cp.clickoncertificateconfirmsave()
        time.sleep(3)
        if "Created and unpublished successfully" in self.driver.page_source:
            self.logger.info("********** certification creation test is passed *********")

        else:
            # Log and take a screenshot
            self.logger.error("************** certification creation test is failed **********")
            self.driver.save_screenshot(".\\Screenshots\\" + "test_certificationcreation.png")
            assert False
        time.sleep(3)
        self.cp.setsearch(certificationname)
        time.sleep(2)
        self.driver.find_element(By.XPATH, "//span[text()='" + certificationname + "']").click()
        self.cp.clickoncertificateedit()
        self.cp.clickoncertificatepublish()
        self.cp.clickoncertificateconfirmpublish()
        time.sleep(3)
        if "Updated and published successfully" in self.driver.page_source:
            self.logger.info("********** certification update test is passed *********")
            self.driver.close()


        else:
            # Log and take a screenshot
            self.logger.error("************** certification update test is failed **********")
            self.driver.save_screenshot(".\\Screenshots\\" + "test_certificateupdate.png")
            assert False
        time.sleep(3)






    @pytest.mark.sanity
    @pytest.mark.regression
    @pytest.mark.run(order=3)
    # @pytest.mark.skip(reason="Skipping this test")
    def test_certificationverificationforrelationcompany(self, setup):
        self.logger.info("************* Test_003_Certification **********")
        self.driver = setup
        self.driver.get(self.baseURL)
        self.driver.maximize_window()

        self.lp = LoginPage(self.driver)
        self.lp.setUserName(self.usernames4)
        self.lp.setPassword(self.password)
        self.lp.clickLogin()
        self.logger.info("************* Login succesful **********")
        self.cp = Certification(self.driver)
        self.cp.clickoncertificationprogramme()
        self.cp.clickonmyexams()
        workbook = load_workbook("TestData/LoginData.xlsx")

        # Access the active worksheet
        worksheet = workbook.active

        certificationname = worksheet["A21"].value

        workbook.close()

        self.cp.setsearch(certificationname)
        time.sleep(3)
        if "Your certification is typically displayed as a document stating that as a professional, you've been trained, educated and are prepared to meet a specific set of criteria for your role." in self.driver.page_source:
            self.logger.info("********** Certification verification test is passed *********")

        else:
            # Log and take a screenshot
            self.logger.error("************** Certification verification test is failed **********")
            self.driver.save_screenshot(".\\Screenshots\\" + "test_certificateverificationforrelationcompany.png")
            assert False
        time.sleep(3)
        self.cp.clickongetcertificate()
        self.cp.clickonexamcheckbox()
        time.sleep(3)
        if "TAKE A TEST" in self.driver.page_source:
            self.logger.info("********** Exam verification test is passed *********")
            self.driver.close()

        else:
            # Log and take a screenshot
            self.logger.error("************** Exam verification test is failed **********")
            self.driver.save_screenshot(".\\Screenshots\\" + "test_examverificationforrelationcompany.png")
            assert False
        time.sleep(3)


    @pytest.mark.sanity
    @pytest.mark.regression
    @pytest.mark.run(order=4)
    # @pytest.mark.skip(reason="Skipping this test")
    def test_certificationverificationforremployee(self, setup):
        self.logger.info("************* Test_004_Certification **********")
        self.driver = setup
        self.driver.get(self.baseURL)
        self.driver.maximize_window()

        self.lp = LoginPage(self.driver)
        self.lp.setUserName(self.usernames5)
        self.lp.setPassword(self.password)
        self.lp.clickLogin()
        self.logger.info("************* Login succesful **********")
        self.cp = Certification(self.driver)
        self.cp.clickoncertificationprogramme()
        workbook = load_workbook("TestData/LoginData.xlsx")

        # Access the active worksheet
        worksheet = workbook.active

        certificationname = worksheet["A21"].value

        workbook.close()

        self.cp.setsearch(certificationname)
        time.sleep(3)
        if "Your certification is typically displayed as a document stating that as a professional, you've been trained, educated and are prepared to meet a specific set of criteria for your role." in self.driver.page_source:
            self.logger.info("********** certification verification test is passed *********")

        else:
            # Log and take a screenshot
            self.logger.error("************** certification verification test is failed **********")
            self.driver.save_screenshot(".\\Screenshots\\" + "test_certificateverificationforemployee.png")
            assert False
        time.sleep(3)
        self.cp.clickongetcertificate()
        self.cp.clickonexamcheckbox()
        time.sleep(3)
        if "TAKE A TEST" in self.driver.page_source:
            self.logger.info("********** Exam verification test is passed *********")
            self.driver.close()

        else:
            # Log and take a screenshot
            self.logger.error("************** Exam verification test is failed **********")
            self.driver.save_screenshot(".\\Screenshots\\" + "test_examverificationforemployee.png")
            assert False
        time.sleep(3)

















