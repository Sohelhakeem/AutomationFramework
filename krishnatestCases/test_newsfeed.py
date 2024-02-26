import os
import unittest

import pytest
import time

from openpyxl.reader.excel import load_workbook
from selenium.webdriver import ActionChains, Keys
from selenium.webdriver.common.by import By
from selenium import webdriver
from pageObjects.LoginPage import LoginPage
from krishnapageObjects.NewsfeedPage import NewsFeed
from utilities.readProperties import ReadConfig
from utilities.customLogger import LogGen


class Test_001_NewsFeed:
    baseURL = ReadConfig.getApplicationURL()
    # username = ReadConfig.getUseremail()
    password = ReadConfig.getPassword()
    whats = ReadConfig.getwhats()
    whatso = ReadConfig.getwhatso()
    # usernames = ReadConfig.getuseremails()
    # usernames1 = ReadConfig.getuseremails1()
    # usernames2 = ReadConfig.getuseremails2()
    # usernames3 = ReadConfig.getuseremails3()
    whatson = ReadConfig.getwhatson()
    oneimage = ReadConfig.getoneimage()
    fiveimages = ReadConfig.getfiveimages()

    workbook = load_workbook("TestData/LoginData.xlsx")

    # Access the active worksheet
    worksheet = workbook.active

    username = worksheet["A2"].value
    usernames = worksheet["E2"].value
    usernames1 = worksheet["I2"].value
    usernames2 = worksheet["I3"].value
    usernames3 = worksheet["H2"].value

    workbook.close()
    # Load the existing workbook
    wb = load_workbook("TestData/LoginData.xlsx")

    # Select the active worksheet
    ws = wb.active

    # Update the existing cells with new data
    ws['A2'] = username
    ws['E2'] = usernames
    ws['I2'] = usernames1
    ws['I3'] = usernames2
    ws['H1'] = usernames3

    # Save the workbook
    wb.save("TestData/LoginData.xlsx")



    relative_one = "Files/one.png"
    absolute_path1 = os.path.abspath(relative_one)
    relative_two = "Files/two.png"
    absolute_path2 = os.path.abspath(relative_two)
    relative_three = "Files/three.png"
    absolute_path3 = os.path.abspath(relative_three)
    relative_four = "Files/four.png"
    absolute_path4 = os.path.abspath(relative_four)
    relative_five = "Files/five.png"
    absolute_path5 = os.path.abspath(relative_five)
    relative_six = "Files/six.jpg"
    absolute_path6 = os.path.abspath(relative_six)
    relative_video1 = "Files/video1.mp4"
    absolute_pathvideo1 = os.path.abspath(relative_video1)
    whatvideo = ReadConfig.getwhatvideo()
    relative_video2 = "Files/video2.mp4"
    absolute_pathvideo2 = os.path.abspath(relative_video2)
    whatvideos = ReadConfig.getwhatvideos()
    whatyoutube = ReadConfig.getwhatyoutube()
    youtubeurl = ReadConfig.getyoutubeurl()
    whatedit = ReadConfig.getwhatedit()
    commenttext = "hiii gud mrng"
    replytext = "gudmorning all"
    commentedittext = "all "

    logger = LogGen.loggen()  # Logger

    # def setUp(self):
    #     self.logger = LogGen.loggen()
    #     self.driver = webdriver.Chrome()
    #     self.driver.maximize_window()
    #     self.logger.info("*******Opening URL*******")
    #     self.driver.get(self.baseURL)
    #
    #
    #
    # def tearDown(self):
    #     self.driver.quit()

    @pytest.mark.krishna
    @pytest.mark.regression
    @pytest.mark.run(order=1)
    @pytest.mark.flaky(reruns=3,reruns_delay=2)
    # @pytest.mark.skip(reason="Skipping this test")
    def test_newsfeedforemployees(self, setup):
        self.logger.info("************* Test_001_NewsFeed **********")
        self.driver = setup
        self.driver.get(self.baseURL)
        self.driver.maximize_window()

        self.lp = LoginPage(self.driver)
        self.lp.setUserName(self.username)
        self.lp.setPassword(self.password)
        self.lp.clickLogin()
        self.logger.info("************* Login succesful **********")

        self.logger.info("******* Starting NewsFeed Test **********")
        self.nf = NewsFeed(self.driver)
        self.nf.clickOnwhat()
        self.nf.setwhats(self.whats)
        self.nf.clickonpost()
        time.sleep(3)
        if "News feed created successfully" in self.driver.page_source:
            self.logger.info("********** NewsFeed test is passed *********")

        else:
            # Log and take a screenshot
            self.logger.error("************** NewsFeed test is failed **********")
            self.driver.save_screenshot(".\\Screenshots\\" + "test_newsfeedcreation.png")
            assert False
        time.sleep(3)
        self.nf.clickonlogout()
        self.logger.info("************* Logout succesful **********")
        self.lp.setUserNames(self.usernames)
        self.lp.setPassword(self.password)
        self.lp.clickLogin()
        time.sleep(3)
        self.logger.info("************* EmpLogin succesful **********")
        if "Hi,gud mrng employees" in self.driver.page_source:
            self.logger.info("********** NewsFeed display is passed *********")
            self.driver.close()
        else:
            # Log and take a screenshot
            self.logger.error("************** NewsFeed display is failed **********")
            self.driver.save_screenshot(".\\Screenshots\\" + "test_newsfeedverification.png")
            assert False

    @pytest.mark.sanity
    @pytest.mark.regression
    @pytest.mark.run(order=2)
    @pytest.mark.flaky(reruns=3, reruns_delay=2)
    # @pytest.mark.skip(reason="Skipping this test")
    def test_newsfeedforempndrel(self, setup):
        self.logger.info("************* Test_002_NewsFeed **********")
        self.driver = setup
        self.driver.get(self.baseURL)
        self.driver.maximize_window()

        self.lp = LoginPage(self.driver)
        self.lp.setUserName(self.username)
        self.lp.setPassword(self.password)
        self.lp.clickLogin()
        self.logger.info("************* Login succesful **********")

        self.logger.info("******* Starting NewsFeed Test **********")
        self.nf = NewsFeed(self.driver)
        self.nf.clickOnwhat()
        self.nf.setwhatso(self.whatso)
        self.nf.clickonpublic()
        self.nf.clickonpost()
        time.sleep(3)
        if "News feed created successfully" in self.driver.page_source:
            self.logger.info("********** NewsFeed test is passed *********")

        else:
            # Log and take a screenshot
            self.logger.error("************** NewsFeed test is failed **********")
            self.driver.save_screenshot(".\\Screenshots\\" + "test_newsfeedcreation.png")
            assert False
        time.sleep(3)
        self.nf.clickonlogout()
        self.logger.info("************* Logout succesful **********")
        self.lp.setUserNames1(self.usernames1)
        self.lp.setPassword(self.password)
        self.lp.clickLogin()
        time.sleep(3)
        self.logger.info("************* relationcompanyLogin succesful **********")
        if "hii,all employees and relation companies" in self.driver.page_source:
            self.logger.info("********** NewsFeed display is passed *********")
            self.driver.close()
        else:
            # Log and take a screenshot
            self.logger.error("************** NewsFeed display is failed **********")
            self.driver.save_screenshot(".\\Screenshots\\" + "test_newsfeedforempndrel.png")
            assert False

    @pytest.mark.test
    @pytest.mark.regression
    @pytest.mark.run(order=3)
    @pytest.mark.flaky(reruns=3, reruns_delay=2)
    # @pytest.mark.skip(reason="Skipping this test")
    def test_newsfeedforpartners(self, setup):
        self.logger.info("************* Test_003_NewsFeed **********")
        self.driver = setup
        self.driver.get(self.baseURL)
        self.driver.maximize_window()

        self.lp = LoginPage(self.driver)
        self.lp.setUserName(self.username)
        self.lp.setPassword(self.password)
        self.lp.clickLogin()
        self.logger.info("************* Login succesful **********")

        self.logger.info("******* Starting NewsFeed Test **********")
        self.nf = NewsFeed(self.driver)
        self.nf.clickOnwhat()
        self.nf.setwhatson(self.whatson)
        self.nf.clickonemp()
        self.nf.clickonpartner()
        self.nf.clickonpost()
        time.sleep(3)
        if "News feed created successfully" in self.driver.page_source:
            self.logger.info("********** NewsFeed test is passed *********")

        else:
            # Log and take a screenshot
            self.logger.error("************** NewsFeed test is failed **********")
            self.driver.save_screenshot(".\\Screenshots\\" + "test_newsfeedcreation.png")
            assert False
        time.sleep(3)
        self.nf.clickonlogout()
        self.logger.info("************* Logout succesful **********")
        self.lp.setUserNames2(self.usernames2)
        self.lp.setPassword(self.password)
        self.lp.clickLogin()
        time.sleep(3)
        self.logger.info("************* partnerLogin succesful **********")
        if "hii,all partners schedule the meeting" in self.driver.page_source:
            self.logger.info("********** NewsFeed display is passed *********")
            self.driver.close()
        else:
            # Log and take a screenshot
            self.logger.error("************** NewsFeed display is failed **********")
            self.driver.save_screenshot(".\\Screenshots\\" + "test_newsfeedforpartners.png")
            assert False

    @pytest.mark.sanity
    @pytest.mark.regression
    @pytest.mark.run(order=4)
    @pytest.mark.flaky(reruns=3, reruns_delay=2)
    # @pytest.mark.skip(reason="Skipping this test")
    def test_newsfeedforarchived(self, setup):
        self.logger.info("************* Test_004_NewsFeed **********")
        self.driver = setup
        self.driver.get(self.baseURL)
        self.driver.maximize_window()

        self.lp = LoginPage(self.driver)
        self.lp.setUserName(self.username)
        self.lp.setPassword(self.password)
        self.lp.clickLogin()
        self.logger.info("************* Login succesful **********")

        self.logger.info("******* Starting NewsFeed Test **********")
        self.nf = NewsFeed(self.driver)
        self.nf.clickOnwhat()
        self.nf.setwhats(self.whats)
        self.nf.clickonstatus()
        self.nf.clickonpost()
        time.sleep(3)
        if "Post created and saved in archived list" in self.driver.page_source:
            self.logger.info("********** NewsFeed test is passed *********")

        else:
            # Log and take a screenshot
            self.logger.error("************** NewsFeed test is failed **********")
            self.driver.save_screenshot(".\\Screenshots\\" + "test_newsfeedcreation.png")
            assert False
        time.sleep(3)
        self.nf.clickonthreedots()
        self.nf.clickonarchive()
        time.sleep(3)
        if "Hi,gud mrng employees" in self.driver.page_source:
            self.logger.info("********** NewsFeed test is passed *********")
            self.driver.close()

        else:
            # Log and take a screenshot
            self.logger.error("************** NewsFeed test is failed **********")
            self.driver.save_screenshot(".\\Screenshots\\" + "test_newsfeedforarchived.png")
            assert False

    @pytest.mark.sanity
    @pytest.mark.regression
    @pytest.mark.run(order=5)
    @pytest.mark.flaky(reruns=3, reruns_delay=2)
    # @pytest.mark.skip(reason="Skipping this test")
    def test_newsfeedwithimage(self, setup):
        self.logger.info("************* Test_005_NewsFeed **********")
        self.driver = setup
        self.driver.get(self.baseURL)
        self.driver.maximize_window()

        self.lp = LoginPage(self.driver)
        self.lp.setUserName(self.username)
        self.lp.setPassword(self.password)
        self.lp.clickLogin()
        self.logger.info("************* Login succesful **********")

        self.logger.info("******* Starting NewsFeed Test **********")
        self.nf = NewsFeed(self.driver)
        self.nf.clickOnwhat()
        self.nf.setoneimage(self.oneimage)
        self.nf.setgallery(self.absolute_path1)
        self.nf.clickonpost()
        time.sleep(3)
        if "News feed created successfully" in self.driver.page_source:
            self.logger.info("********** NewsFeed test is passed *********")

        else:
            # Log and take a screenshot
            self.logger.error("************** NewsFeed test is failed **********")
            self.driver.save_screenshot(".\\Screenshots\\" + "test_newsfeedcreation.png")
            assert False
        self.nf.clickonlogout()
        self.logger.info("************* Logout succesful **********")
        self.lp.setUserNames(self.usernames)
        self.lp.setPassword(self.password)
        self.lp.clickLogin()
        self.logger.info("************* EmpLogin succesful **********")
        time.sleep(3)
        if "one image upload" in self.driver.page_source:
            self.logger.info("***************Newsfeed test is passed **********")
            self.driver.close()

        else:
            self.logger.error("************* NewsFeed test is failed **********")
            self.driver.save_screenshot(".\\Screenshots\\" + "test_newsfeedwithimage.png")
            assert False

    @pytest.mark.sanity
    @pytest.mark.regression
    @pytest.mark.run(order=6)
    @pytest.mark.flaky(reruns=3, reruns_delay=2)
    # @pytest.mark.skip(reason="Skipping this test")
    def test_newsfeedwith5images(self, setup):
        self.logger.info("************* Test_006_NewsFeed **********")
        self.driver = setup
        self.driver.get(self.baseURL)
        self.driver.maximize_window()

        self.lp = LoginPage(self.driver)
        self.lp.setUserName(self.username)
        self.lp.setPassword(self.password)
        self.lp.clickLogin()
        self.logger.info("************* Login succesful **********")

        self.logger.info("******* Starting NewsFeed Test **********")
        self.nf = NewsFeed(self.driver)
        self.nf.clickOnwhat()
        self.nf.setfiveimages(self.fiveimages)
        self.nf.setgallery(self.absolute_path1)
        self.nf.setgallery(self.absolute_path2)
        self.nf.setgallery(self.absolute_path3)
        self.nf.setgallery(self.absolute_path4)
        self.nf.setgallery(self.absolute_path5)
        self.nf.clickonpost()
        time.sleep(3)
        if "News feed created successfully" in self.driver.page_source:
            self.logger.info("********** NewsFeed test is passed *********")

        else:
            # Log and take a screenshot
            self.logger.error("************** NewsFeed test is failed **********")
            self.driver.save_screenshot(".\\Screenshots\\" + "test_newsfeedcreation.png")
            assert False
        self.nf.clickonlogout()
        self.logger.info("************* Logout succesful **********")
        self.lp.setUserNames(self.usernames)
        self.lp.setPassword(self.password)
        self.lp.clickLogin()
        self.logger.info("************* EmpLogin succesful **********")
        self.nf.clickonimage()
        time.sleep(3)
        if "five images upload" in self.driver.page_source:
            self.logger.info("***************Newsfeed test is passed **********")
            self.driver.close()

        else:
            self.logger.error("************* NewsFeed test is failed **********")
            self.driver.save_screenshot(".\\Screenshots\\" + "test_newsfeedwith5images.png")
            assert False

    @pytest.mark.sanity
    @pytest.mark.regression
    @pytest.mark.run(order=7)
    @pytest.mark.flaky(reruns=3, reruns_delay=2)
    # @pytest.mark.skip(reason="Skipping this test")
    def test_newsfeedwith6images(self, setup):
        self.logger.info("************* Test_007_NewsFeed **********")
        self.driver = setup
        self.driver.get(self.baseURL)
        self.driver.maximize_window()

        self.lp = LoginPage(self.driver)
        self.lp.setUserName(self.username)
        self.lp.setPassword(self.password)
        self.lp.clickLogin()
        self.logger.info("************* Login succesful **********")

        self.logger.info("******* Starting NewsFeed Test **********")
        self.nf = NewsFeed(self.driver)
        self.nf.clickOnwhat()
        self.nf.setgallery(self.absolute_path1)
        self.nf.setgallery(self.absolute_path2)
        self.nf.setgallery(self.absolute_path3)
        self.nf.setgallery(self.absolute_path4)
        self.nf.setgallery(self.absolute_path5)
        self.nf.setgallery(self.absolute_path6)
        # self.nf.clickonpost()
        time.sleep(3)
        if "Cannot upload more than 5 images" in self.driver.page_source:
            self.logger.info("********** NewsFeed test is passed *********")
            self.driver.close()

        else:
            # Log and take a screenshot
            self.logger.error("************** NewsFeed test is failed **********")
            self.driver.save_screenshot(".\\Screenshots\\" + "test_newsfeedwith6images.png")
            assert False

    @pytest.mark.sanity
    @pytest.mark.regression
    @pytest.mark.run(order=8)
    @pytest.mark.flaky(reruns=3, reruns_delay=2)
    # @pytest.mark.skip(reason="Skipping this test")
    def test_newsfeedwithimages(self, setup):
        self.logger.info("************* Test_008_NewsFeed **********")
        self.driver = setup
        self.driver.get(self.baseURL)
        self.driver.maximize_window()

        self.lp = LoginPage(self.driver)
        self.lp.setUserName(self.username)
        self.lp.setPassword(self.password)
        self.lp.clickLogin()
        self.logger.info("************* Login succesful **********")

        self.logger.info("******* Starting NewsFeed Test **********")
        self.nf = NewsFeed(self.driver)
        self.nf.clickOnwhat()
        self.nf.setfiveimages(self.fiveimages)
        self.nf.setgallery(self.absolute_path1)
        self.nf.setgallery(self.absolute_path2)
        self.nf.setgallery(self.absolute_path3)
        self.nf.setgallery(self.absolute_path4)
        self.nf.setgallery(self.absolute_path5)
        self.nf.clickonpublic()
        self.nf.clickonpost()
        time.sleep(3)
        if "News feed created successfully" in self.driver.page_source:
            self.logger.info("********** NewsFeed test is passed *********")

        else:
            # Log and take a screenshot
            self.logger.error("************** NewsFeed test is failed **********")
            self.driver.save_screenshot(".\\Screenshots\\" + "test_newsfeedcreation.png")
            assert False
        self.nf.clickonlogout()
        self.logger.info("************* Logout succesful **********")
        self.lp.setUserNames1(self.usernames1)
        self.lp.setPassword(self.password)
        self.lp.clickLogin()
        self.logger.info("************* EmpLogin succesful **********")
        self.nf.clickonimage()
        time.sleep(3)
        if "five images upload" in self.driver.page_source:
            self.logger.info("***************Newsfeed test is passed **********")
            self.driver.close()

        else:
            self.logger.error("************* NewsFeed test is failed **********")
            self.driver.save_screenshot(".\\Screenshots\\" + "test_newsfeedwith6images.png")
            assert False

    @pytest.mark.sanity
    @pytest.mark.regression
    @pytest.mark.run(order=9)
    @pytest.mark.flaky(reruns=3, reruns_delay=2)
    # @pytest.mark.skip(reason="Skipping this test")

    def test_newsfeedforemployeesvideo(self, setup):
        self.logger.info("************* Test_009_NewsFeed **********")
        self.driver = setup
        self.driver.get(self.baseURL)
        self.driver.maximize_window()

        self.lp = LoginPage(self.driver)
        self.lp.setUserName(self.username)
        self.lp.setPassword(self.password)
        self.lp.clickLogin()
        self.logger.info("************* Login succesful **********")

        self.logger.info("******* Starting NewsFeed Test **********")
        self.nf = NewsFeed(self.driver)
        self.nf.clickOnwhat()
        self.nf.setwhats(self.whatvideo)
        self.nf.setvideo(self.absolute_pathvideo1)
        time.sleep(3)
        # self.nf.clickonpost()
        if "Cannot upload greater than 20MB" in self.driver.page_source:
            self.logger.info("********** NewsFeed test is passed *********")
            self.driver.close()

        else:
            # Log and take a screenshot
            self.logger.error("************** NewsFeed test is failed **********")
            self.driver.save_screenshot(".\\Screenshots\\" + "test_newsfeedforemployeesvideo.png")
            assert False

    @pytest.mark.sanity
    @pytest.mark.regression
    @pytest.mark.run(order=10)
    @pytest.mark.flaky(reruns=3, reruns_delay=2)
    # @pytest.mark.skip(reason="Skipping this test")
    def test_newsfeedforemployeesvideos(self, setup):
        self.logger.info("************* Test_010_NewsFeed **********")
        self.driver = setup
        self.driver.get(self.baseURL)
        self.driver.maximize_window()

        self.lp = LoginPage(self.driver)
        self.lp.setUserName(self.username)
        self.lp.setPassword(self.password)
        self.lp.clickLogin()
        self.logger.info("************* Login succesful **********")

        self.logger.info("******* Starting NewsFeed Test **********")
        self.nf = NewsFeed(self.driver)
        self.nf.clickOnwhat()
        self.nf.setwhats(self.whatvideos)
        self.nf.setvideo(self.absolute_pathvideo2)
        self.nf.clickonpost()
        time.sleep(3)
        if "News feed created successfully" in self.driver.page_source:
            self.logger.info("********** NewsFeed test is passed *********")

        else:
            # Log and take a screenshot
            self.logger.error("************** NewsFeed test is failed **********")
            self.driver.save_screenshot(".\\Screenshots\\" + "test_newsfeedcreation.png")
            assert False
        time.sleep(3)
        self.nf.clickonlogout()
        self.logger.info("************* Logout succesful **********")
        self.lp.setUserNames(self.usernames)
        self.lp.setPassword(self.password)
        self.lp.clickLogin()
        time.sleep(3)
        self.logger.info("************* EmpLogin succesful **********")
        if "uploading videos" in self.driver.page_source:
            self.logger.info("********** NewsFeed display is passed *********")
            self.driver.close()
        else:
            # Log and take a screenshot
            self.logger.error("************** NewsFeed display is failed **********")
            self.driver.save_screenshot(".\\Screenshots\\" + "test_newsfeedforemployeesvideos.png")
            assert False

    @pytest.mark.sanity
    @pytest.mark.regression
    @pytest.mark.run(order=11)
    @pytest.mark.flaky(reruns=3, reruns_delay=2)
    # @pytest.mark.skip(reason="Skipping this test")
    def test_newsfeedforempndrelvideo(self, setup):
        self.logger.info("************* Test_011_NewsFeed **********")
        self.driver = setup
        self.driver.get(self.baseURL)
        self.driver.maximize_window()

        self.lp = LoginPage(self.driver)
        self.lp.setUserName(self.username)
        self.lp.setPassword(self.password)
        self.lp.clickLogin()
        self.logger.info("************* Login succesful **********")

        self.logger.info("******* Starting NewsFeed Test **********")
        self.nf = NewsFeed(self.driver)
        self.nf.clickOnwhat()
        self.nf.setwhatyoutube(self.whatyoutube)
        self.nf.clickonpublic()
        self.nf.clickonyoutube()
        self.nf.setyoutubeurl(self.youtubeurl)
        self.nf.clickondone()
        self.nf.clickonpost()
        time.sleep(3)
        if "News feed created successfully" in self.driver.page_source:
            self.logger.info("********** NewsFeed test is passed *********")

        else:
            # Log and take a screenshot
            self.logger.error("************** NewsFeed test is failed **********")
            self.driver.save_screenshot(".\\Screenshots\\" + "test_newsfeedcreation.png")
            assert False
        time.sleep(3)
        self.nf.clickonlogout()
        self.logger.info("************* Logout succesful **********")
        self.lp.setUserNames1(self.usernames1)
        self.lp.setPassword(self.password)
        self.lp.clickLogin()
        time.sleep(3)
        self.logger.info("************* relationcompanyLogin succesful **********")
        if "uploading youtube videos" in self.driver.page_source:
            self.logger.info("********** NewsFeed display is passed *********")
            self.driver.close()
        else:
            # Log and take a screenshot
            self.logger.error("************** NewsFeed display is failed **********")
            self.driver.save_screenshot(".\\Screenshots\\" + "test_newsfeedforempndrelvideo.png")
            assert False

    @pytest.mark.sanity
    @pytest.mark.regression
    @pytest.mark.run(order=12)
    @pytest.mark.flaky(reruns=3, reruns_delay=2)
    # @pytest.mark.skip(reason="Skipping this test")
    def test_editnewsfeedforemployees(self, setup):
        self.logger.info("************* Test_012_NewsFeed **********")
        self.driver = setup
        self.driver.get(self.baseURL)
        self.driver.maximize_window()

        self.lp = LoginPage(self.driver)
        self.lp.setUserName(self.username)
        self.lp.setPassword(self.password)
        self.lp.clickLogin()
        self.logger.info("************* Login succesful **********")

        self.logger.info("******* Starting NewsFeed Test **********")
        self.nf = NewsFeed(self.driver)
        self.nf.clickOnwhat()
        self.nf.setwhats(self.whats)
        self.nf.clickonpost()
        time.sleep(3)
        if "News feed created successfully" in self.driver.page_source:
            self.logger.info("********** NewsFeed test is passed *********")

        else:
            # Log and take a screenshot
            self.logger.error("************** NewsFeed test is failed **********")
            self.driver.save_screenshot(".\\Screenshots\\" + "test_newsfeedcreation.png")
            assert False
        time.sleep(3)
        self.nf.clickonthreedot()
        self.nf.clickonedit()
        self.nf.setwhatedit(self.whatedit)
        self.nf.clickonpublic()
        self.nf.clickonupdate()
        time.sleep(3)
        if "News feed updated successfully" in self.driver.page_source:
            self.logger.info("********** NewsFeed test is passed *********")

        else:
            # Log and take a screenshot
            self.logger.error("************** NewsFeed test is failed **********")
            self.driver.save_screenshot(".\\Screenshots\\" + "test_newsfeedupdate.png")
            assert False
        time.sleep(3)
        self.nf.clickonlogout()
        self.logger.info("************* Logout succesful **********")
        self.lp.setUserNames1(self.usernames1)
        self.lp.setPassword(self.password)
        self.lp.clickLogin()
        time.sleep(3)
        self.logger.info("************* relationcompanyLogin succesful **********")
        if "Hi,gud mrng employeesEditing the feed" in self.driver.page_source:
            self.logger.info("********** NewsFeed display is passed *********")
            self.driver.close()
        else:
            # Log and take a screenshot
            self.logger.error("************** NewsFeed display is failed **********")
            self.driver.save_screenshot(".\\Screenshots\\" + "test_editnewsfeedforemployees.png")
            assert False

    @pytest.mark.sanity
    @pytest.mark.regression
    @pytest.mark.run(order=13)
    @pytest.mark.flaky(reruns=3, reruns_delay=2)
    # @pytest.mark.skip(reason="Skipping this test")
    def test_deletenewsfeedforempndrelvideo(self, setup):
        self.logger.info("************* Test_013_NewsFeed **********")
        self.driver = setup
        self.driver.get(self.baseURL)
        self.driver.maximize_window()

        self.lp = LoginPage(self.driver)
        self.lp.setUserName(self.username)
        self.lp.setPassword(self.password)
        self.lp.clickLogin()
        self.logger.info("************* Login succesful **********")

        self.logger.info("******* Starting NewsFeed Test **********")
        self.nf = NewsFeed(self.driver)
        self.nf.clickOnwhat()
        self.nf.setwhatyoutube(self.whatyoutube)
        self.nf.clickonpublic()
        self.nf.clickonyoutube()
        self.nf.setyoutubeurl(self.youtubeurl)
        self.nf.clickondone()
        self.nf.clickonpost()
        time.sleep(3)
        if "News feed created successfully" in self.driver.page_source:
            self.logger.info("********** NewsFeed test is passed *********")

        else:
            # Log and take a screenshot
            self.logger.error("************** NewsFeed test is failed **********")
            self.driver.save_screenshot(".\\Screenshots\\" + "test_newsfeedcreation.png")
            assert False
        time.sleep(3)
        self.nf.clickonthreedot()
        self.nf.clickondelete()
        self.nf.clickondeletecheckbox()
        self.nf.clickondeletebutton()
        time.sleep(3)
        if "News Feed deleted successfully!" in self.driver.page_source:
            self.logger.info("********** NewsFeed test is passed *********")
            self.driver.close()

        else:
            # Log and take a screenshot
            self.logger.error("************** NewsFeed test is failed **********")
            self.driver.save_screenshot(".\\Screenshots\\" + "test_deletenewsfeedforempndrelvideo.png")
            assert False

    @pytest.mark.sanity
    @pytest.mark.regression
    @pytest.mark.run(order=14)
    @pytest.mark.flaky(reruns=3, reruns_delay=2)
    # @pytest.mark.skip(reason="Skipping this test")
    def test_archivenewsfeedwith5images(self, setup):
        self.logger.info("************* Test_014_NewsFeed **********")
        self.driver = setup
        self.driver.get(self.baseURL)
        self.driver.maximize_window()

        self.lp = LoginPage(self.driver)
        self.lp.setUserName(self.username)
        self.lp.setPassword(self.password)
        self.lp.clickLogin()
        self.logger.info("************* Login succesful **********")

        self.logger.info("******* Starting NewsFeed Test **********")
        self.nf = NewsFeed(self.driver)
        self.nf.clickOnwhat()
        self.nf.setfiveimages(self.fiveimages)
        self.nf.setgallery(self.absolute_path1)
        self.nf.setgallery(self.absolute_path2)
        self.nf.setgallery(self.absolute_path3)
        self.nf.setgallery(self.absolute_path4)
        self.nf.setgallery(self.absolute_path5)
        self.nf.clickonpost()
        time.sleep(3)
        if "News feed created successfully" in self.driver.page_source:
            self.logger.info("********** NewsFeed test is passed *********")

        else:
            # Log and take a screenshot
            self.logger.error("************** NewsFeed test is failed **********")
            self.driver.save_screenshot(".\\Screenshots\\" + "test_newsfeedcreation.png")
            assert False
        time.sleep(3)
        self.nf.clickonthreedot()
        self.nf.clickondelete()
        self.nf.clickondeletebutton()
        time.sleep(3)
        if "Feed deleted successfully, saved in archived feed" in self.driver.page_source:
            self.logger.info("********** NewsFeed test is passed *********")

        else:
            # Log and take a screenshot
            self.logger.error("************** NewsFeed test is failed **********")
            self.driver.save_screenshot(".\\Screenshots\\" + "test_newsfeedarchieve.png")
            assert False
        time.sleep(3)
        self.nf.clickonthreedots()
        self.nf.clickonarchive()
        time.sleep(3)
        if "five images upload" in self.driver.page_source:
            self.logger.info("********** NewsFeed test is passed *********")
            self.driver.close()

        else:
            # Log and take a screenshot
            self.logger.error("************** NewsFeed test is failed **********")
            self.driver.save_screenshot(".\\Screenshots\\" + "test_archivenewsfeedwith5images.png")
            assert False

    @pytest.mark.sanity
    @pytest.mark.regression
    @pytest.mark.run(order=15)
    @pytest.mark.flaky(reruns=3, reruns_delay=2)
    # @pytest.mark.skip(reason="Skipping this test")
    def test_bookmarknewsfeed(self, setup):
        self.logger.info("************* Test_015_NewsFeed **********")
        self.driver = setup
        self.driver.get(self.baseURL)
        self.driver.maximize_window()

        self.lp = LoginPage(self.driver)
        self.lp.setUserName(self.username)
        self.lp.setPassword(self.password)
        self.lp.clickLogin()
        self.logger.info("************* companyLogin succesful **********")
        self.logger.info("******* Starting NewsFeed Test **********")
        self.nf = NewsFeed(self.driver)
        self.nf.clickOnwhat()
        self.nf.setwhatso(self.whatso)
        self.nf.clickonpublic()
        self.nf.clickonpost()
        time.sleep(3)
        if "News feed created successfully" in self.driver.page_source:
            self.logger.info("********** NewsFeed test is passed *********")

        else:
            # Log and take a screenshot
            self.logger.error("************** NewsFeed test is failed **********")
            self.driver.save_screenshot(".\\Screenshots\\" + "test_newsfeedcreation.png")
            assert False
        time.sleep(3)
        self.nf = NewsFeed(self.driver)
        self.nf.clickonthreedot()
        self.nf.clickonbookmark()
        self.nf.clickonexplore()
        time.sleep(3)
        if "hii,all employees and relation companies" in self.driver.page_source:
            self.logger.info("********** NewsFeed test is passed *********")

        else:
            # Log and take a screenshot
            self.logger.error("************** NewsFeed test is failed **********")
            self.driver.save_screenshot(".\\Screenshots\\" + "test_newsfeedverification.png")
            assert False
        time.sleep(3)
        self.nf.clickonremove()
        time.sleep(3)
        if "hii,all employees and relation companies" in self.driver.page_source:
            self.logger.error("********** NewsFeed test is failed *********")
            # Log and take a screenshot
            self.driver.save_screenshot(".\\Screenshots\\" + "test_bookmarknewsfeed.png")
            assert False


        else:
            self.logger.info("************** NewsFeed test is passed **********")
            self.driver.close()
        time.sleep(3)

    @pytest.mark.sanity
    @pytest.mark.regression
    @pytest.mark.run(order=16)
    @pytest.mark.flaky(reruns=3, reruns_delay=2)
    # @pytest.mark.skip(reason="Skipping this test")
    def test_selfposts(self, setup):
        self.logger.info("************* Test_016_NewsFeed **********")
        self.driver = setup
        self.driver.get(self.baseURL)
        self.driver.maximize_window()

        self.lp = LoginPage(self.driver)
        self.lp.setUserName(self.username)
        self.lp.setPassword(self.password)
        self.lp.clickLogin()
        self.logger.info("************* companyLogin succesful **********")
        self.logger.info("******* Starting NewsFeed Test **********")
        self.nf = NewsFeed(self.driver)
        self.nf.clickOnwhat()
        self.nf.setwhatso(self.whatso)
        self.nf.clickonpublic()
        self.nf.clickonpost()
        time.sleep(3)
        if "News feed created successfully" in self.driver.page_source:
            self.logger.info("********** NewsFeed test is passed *********")

        else:
            # Log and take a screenshot
            self.logger.error("************** NewsFeed test is failed **********")
            self.driver.save_screenshot(".\\Screenshots\\" + "test_newsfeedcreation.png")
            assert False
        time.sleep(3)
        self.nf.clickonthreedots()
        self.nf.clickonfilter()
        self.nf.clickonall()
        self.nf.clickonself()
        self.nf.clickonapply()
        time.sleep(3)
        if "hii,all employees and relation companies" in self.driver.page_source:
            self.logger.info("********** NewsFeed display is passed *********")
            self.driver.close()
        else:
            # Log and take a screenshot
            self.logger.error("************** NewsFeed display is failed **********")
            self.driver.save_screenshot(".\\Screenshots\\" + "test_selfposts.png")
            assert False

    @pytest.mark.sanity
    @pytest.mark.regression
    @pytest.mark.run(order=17)
    @pytest.mark.flaky(reruns=3, reruns_delay=2)
    # @pytest.mark.skip(reason="Skipping this test")
    def test_sharenewsfeed(self, setup):
        self.logger.info("************* Test_017_NewsFeed **********")
        self.driver = setup
        self.driver.get(self.baseURL)
        self.driver.maximize_window()

        self.lp = LoginPage(self.driver)
        self.lp.setUserName(self.username)
        self.lp.setPassword(self.password)
        self.lp.clickLogin()
        self.logger.info("************* companyLogin succesful **********")
        self.logger.info("******* Starting NewsFeed Test **********")
        self.nf = NewsFeed(self.driver)
        self.nf.clickOnwhat()
        self.nf.setwhatso(self.whatso)
        self.nf.clickonpublic()
        self.nf.clickonpost()
        time.sleep(1)
        if "News feed created successfully" in self.driver.page_source:
            self.logger.info("********** NewsFeed test is passed *********")

        else:
            # Log and take a screenshot
            self.logger.error("************** NewsFeed test is failed **********")
            self.driver.save_screenshot(".\\Screenshots\\" + "test_newsfeedcreation.png")
            assert False
        time.sleep(1)
        self.nf.clickonshare()
        self.nf.clickonwhatsapp()
        window_handles = self.driver.window_handles
        self.driver.switch_to.window(window_handles[1])
        time.sleep(3)

        if "https://web.whatsapp.com/" in self.driver.current_url:
            self.logger.info("********** NewsFeed test is passed *********")

        else:
            # Log and take a screenshot
            self.logger.error("************** NewsFeed test is failed **********")
            self.driver.save_screenshot(".\\Screenshots\\" + "test_sharenewsfeedwhatsapp.png")
            assert False
        time.sleep(3)
        window_handles = self.driver.window_handles
        # Switch to the last opened window
        self.driver.switch_to.window(window_handles[0])
        self.nf.clickonshare()
        self.driver.implicitly_wait(5)
        self.nf.clickonfacebook()
        window_handles = self.driver.window_handles
        self.driver.switch_to.window(window_handles[2])
        time.sleep(3)

        if "https://www.facebook.com/" in self.driver.current_url:
            self.logger.info("********** NewsFeed test is passed *********")

        else:
            # Log and take a screenshot
            self.logger.error("************** NewsFeed test is failed **********")
            self.driver.save_screenshot(".\\Screenshots\\" + "test_sharenewsfeedfacebook.png")
            assert False
        time.sleep(3)
        window_handles = self.driver.window_handles
        self.driver.switch_to.window(window_handles[0])
        self.nf.clickonshare()
        self.nf.clickontwitter()
        window_handles = self.driver.window_handles
        self.driver.switch_to.window(window_handles[3])
        time.sleep(3)

        if "https://twitter.com/" in self.driver.current_url:
            self.logger.info("********** NewsFeed test is passed *********")

        else:
            # Log and take a screenshot
            self.logger.error("************** NewsFeed test is failed **********")
            self.driver.save_screenshot(".\\Screenshots\\" + "test_sharenewsfeedtwitter.png")
            assert False
        time.sleep(3)
        window_handles = self.driver.window_handles
        self.driver.switch_to.window(window_handles[0])
        self.nf.clickonshare()
        self.nf.clickonlinkedin()
        window_handles = self.driver.window_handles
        self.driver.switch_to.window(window_handles[4])
        time.sleep(3)

        if "https://www.linkedin.com/" in self.driver.current_url:
            self.logger.info("********** NewsFeed test is passed *********")

        else:
            # Log and take a screenshot
            self.logger.error("************** NewsFeed test is failed **********")
            self.driver.save_screenshot(".\\Screenshots\\" + "test_sharenewsfeedlinkedin.png")
            assert False
        time.sleep(3)
        window_handles = self.driver.window_handles
        self.driver.switch_to.window(window_handles[0])
        self.nf.clickonshare()
        self.nf.clickontelegram()
        window_handles = self.driver.window_handles
        self.driver.switch_to.window(window_handles[5])
        time.sleep(3)

        if "https://telegram.me/" in self.driver.current_url:
            self.logger.info("********** NewsFeed test is passed *********")

        else:
            # Log and take a screenshot
            self.logger.error("************** NewsFeed test is failed **********")
            self.driver.save_screenshot(".\\Screenshots\\" + "test_sharenewsfeedtelegram.png")
            assert False
        time.sleep(3)
        window_handles = self.driver.window_handles
        self.driver.switch_to.window(window_handles[0])
        self.nf.clickonshare()
        self.nf.clickongmail()
        window_handles = self.driver.window_handles
        self.driver.switch_to.window(window_handles[6])
        time.sleep(3)

        if "https://accounts.google.com/" in self.driver.current_url:
            self.logger.info("********** NewsFeed test is passed *********")
        elif "https://www.google.com/" in self.driver.current_url:
            # Handle the other condition
            self.logger.info("********** NewsFeed test is passed *********")

        else:
            # Log and take a screenshot
            self.logger.error("************** NewsFeed test is failed **********")
            self.driver.save_screenshot(".\\Screenshots\\" + "test_sharenewsfeedgoogle.png")
            assert False
        time.sleep(3)
        window_handles = self.driver.window_handles
        self.driver.switch_to.window(window_handles[0])
        # self.nf.clickonshare()
        # time.sleep(2)
        self.nf.clickoninstagram()
        window_handles = self.driver.window_handles
        self.driver.switch_to.window(window_handles[7])
        time.sleep(3)

        if "https://www.instagram.com/" in self.driver.current_url:
            self.logger.info("********** NewsFeed test is passed *********")
            self.driver.close()

        else:
            # Log and take a screenshot
            self.logger.error("************** NewsFeed test is failed **********")
            self.driver.save_screenshot(".\\Screenshots\\" + "test_sharenewsfeedinstagram.png")
            assert False

    @pytest.mark.sanity
    @pytest.mark.regression
    @pytest.mark.run(order=18)
    @pytest.mark.flaky(reruns=3, reruns_delay=2)
    # @pytest.mark.skip(reason="Skipping this test")
    def test_adminfeed(self, setup):
        self.logger.info("************* Test_018_NewsFeed **********")
        self.driver = setup
        self.driver.get(self.baseURL)
        self.driver.maximize_window()

        self.lp = LoginPage(self.driver)
        self.lp.setUserNames3(self.usernames3)
        self.lp.setPassword(self.password)
        self.lp.clickLogin()
        self.logger.info("************* Login succesful **********")

        self.logger.info("******* Starting NewsFeed Test **********")
        self.nf = NewsFeed(self.driver)
        self.nf.clickOnwhat()
        self.nf.setwhatso(self.whatso)
        self.nf.clickonpublic()
        self.nf.clickonpost()
        time.sleep(3)
        if "News feed created successfully" in self.driver.page_source:
            self.logger.info("********** NewsFeed test is passed *********")

        else:
            # Log and take a screenshot
            self.logger.error("************** NewsFeed test is failed **********")
            self.driver.save_screenshot(".\\Screenshots\\" + "test_newsfeedcreation.png")
            assert False
        time.sleep(3)
        self.nf.clickonlogout()
        self.logger.info("************* Logout succesful **********")
        self.lp.setUserNames1(self.usernames1)
        self.lp.setPassword(self.password)
        self.lp.clickLogin()
        time.sleep(3)
        self.logger.info("************* relationcompanyLogin succesful **********")
        if "hii,all employees and relation companies" in self.driver.page_source:
            self.logger.info("********** NewsFeed display is passed *********")
            self.driver.close()
        else:
            # Log and take a screenshot
            self.logger.error("************** NewsFeed display is failed **********")
            self.driver.save_screenshot(".\\Screenshots\\" + "test_adminfeed.png")
            assert False

    @pytest.mark.sanity
    @pytest.mark.regression
    @pytest.mark.run(order=19)
    @pytest.mark.flaky(reruns=3, reruns_delay=2)
    # @pytest.mark.skip(reason="Skipping this test")
    def test_feedcomment(self, setup):
        self.logger.info("************* Test_019_NewsFeed **********")
        self.driver = setup
        self.driver.get(self.baseURL)
        self.driver.maximize_window()

        self.lp = LoginPage(self.driver)
        self.lp.setUserName(self.username)
        self.lp.setPassword(self.password)
        self.lp.clickLogin()
        self.logger.info("************* Login succesful **********")

        self.logger.info("******* Starting NewsFeed Test **********")
        self.nf = NewsFeed(self.driver)
        self.nf.clickOnwhat()
        self.nf.setwhats(self.whats)
        self.nf.clickonpost()
        time.sleep(3)
        if "News feed created successfully" in self.driver.page_source:
            self.logger.info("********** NewsFeed test is passed *********")

        else:
            # Log and take a screenshot
            self.logger.error("************** NewsFeed test is failed **********")
            self.driver.save_screenshot(".\\Screenshots\\" + "test_newsfeedcreation.png")
            assert False
        time.sleep(3)
        self.nf.clickoncomment()
        self.nf.setcommenttext(self.commenttext)
        self.nf.clickonsend()
        self.nf.clickoncancelcmnt()
        self.nf.clickonlogout()
        self.logger.info("************* Logout succesful **********")
        self.driver.execute_script("window.open('', '_blank');")
        window_handles = self.driver.window_handles
        self.driver.switch_to.window(window_handles[1])
        self.driver.get(self.baseURL)
        self.lp.setUserNames(self.usernames)
        self.lp.setPassword(self.password)
        self.lp.clickLogin()
        time.sleep(3)
        self.logger.info("************* EmpLogin succesful **********")
        if "Hi,gud mrng employees" in self.driver.page_source:
            self.logger.info("********** NewsFeed display is passed *********")

        else:
            # Log and take a screenshot
            self.logger.error("************** NewsFeed display is failed **********")
            self.driver.save_screenshot(".\\Screenshots\\" + "test_newsfeedverification.png")
            assert False
        time.sleep(3)
        self.nf.clickoncomments()
        self.nf.clickonreplycmnt()
        self.nf.setreplytext(self.replytext)
        self.nf.clickonreplysend()
        self.nf.clickonviewreply()
        time.sleep(3)
        if "gudmorning all" in self.driver.page_source:
            self.logger.info("********** NewsFeed display is passed *********")
            self.driver.close()

        else:
            # Log and take a screenshot
            self.logger.error("************** NewsFeed display is failed **********")
            self.driver.save_screenshot(".\\Screenshots\\" + "test_feedcomment.png")
            assert False

    @pytest.mark.sanity
    @pytest.mark.regression
    @pytest.mark.run(order=20)
    @pytest.mark.flaky(reruns=3, reruns_delay=2)
    # @pytest.mark.skip(reason="Skipping this test")
    def test_editfeedcomment(self, setup):
        self.logger.info("************* Test_020_NewsFeed **********")
        self.driver = setup
        self.driver.get(self.baseURL)
        self.driver.maximize_window()

        self.lp = LoginPage(self.driver)
        self.lp.setUserName(self.username)
        self.lp.setPassword(self.password)
        self.lp.clickLogin()
        self.logger.info("************* Login succesful **********")

        self.logger.info("******* Starting NewsFeed Test **********")
        self.nf = NewsFeed(self.driver)
        self.nf.clickOnwhat()
        self.nf.setwhats(self.whats)
        self.nf.clickonpost()
        time.sleep(3)
        if "News feed created successfully" in self.driver.page_source:
            self.logger.info("********** NewsFeed test is passed *********")

        else:
            # Log and take a screenshot
            self.logger.error("************** NewsFeed test is failed **********")
            self.driver.save_screenshot(".\\Screenshots\\" + "test_newsfeedcreation.png")
            assert False
        time.sleep(3)
        self.nf.clickoncomment()
        self.nf.setcommenttext(self.commenttext)
        self.nf.clickonsend()
        self.nf.clickoncommentthreedot()
        self.nf.clickoncommentedit()
        self.nf.setcommentedittext(self.commentedittext)
        self.nf.clickoncommenteditsend()
        self.nf.clickoncommentthreedot()
        self.nf.clickoncommentdelete()
        self.nf.clickoncommentconfirmdelete()
        time.sleep(3)
        if "all hiii gud mrng" in self.driver.page_source:
            self.logger.error("********** NewsFeed test is failed *********")
            # Log and take a screenshot
            self.driver.save_screenshot(".\\Screenshots\\" + "test_editfeedcomment.png")
            assert False


        else:
            self.logger.info("************** NewsFeed test is passed **********")
            self.driver.close()

    # if __name__ == '__main__':
    #     unittest.main(verbosity=2)
